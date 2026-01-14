import argparse
import csv
import json
import logging
import os
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from typing import Optional, Tuple, Any, Dict, List

# SprÃ³buj zaimportowaÄ‡ biblioteki, jeÅ›li ich nie ma, poinformuj uÅ¼ytkownika
try:
    import pandas as pd
    from tqdm import tqdm
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
except ImportError:
    print(
        "BÅ‚Ä…d: Wymagane biblioteki (pandas, openpyxl, tqdm) nie sÄ… zainstalowane.",
        file=sys.stderr,
    )
    print(
        "ProszÄ™ je zainstalowaÄ‡ komendÄ…: pip install pandas openpyxl tqdm",
        file=sys.stderr,
    )
    sys.exit(1)

from refresh_zoho_access_token import refresh_access_token

# --- Konfiguracja API Zoho (wspÃ³lna z get_zoho_account_details.py) ---
ZOHO_CRM_API_BASE_URL = "https://www.zohoapis.eu/crm/v3"
ZOHO_CRM_COQL_URL = f"{ZOHO_CRM_API_BASE_URL}/coql"
REQUEST_SLEEP_S = 0.2  # delikatny break miÄ™dzy wywoÅ‚aniami API
MAX_RETRY_TIMEOUT = 2  # liczba ponowieÅ„ przy timeoutach sieciowych

# Aktualny token (odÅ›wieÅ¼any automatycznie przy 401/403)
ACCESS_TOKEN_HOLDER: Dict[str, Any] = {"token": None, "refreshed_at": 0.0}

# --- Definicje statusÃ³w ---
AKTYWNE_LEAD_STATUSY = [
    "Aktywny zimny", "Dzwonienie", "Nurturing",
    "UmÃ³wione spotkanie", "Zakwalifikowane do sales"
]

AKTYWNE_DEAL_ETAPY = [
    "Qualification", "Spotkania", "Oferta", "Negocjacje",
    "Oferta zaakceptowana", "Umowa w podpisie"
]

WYGRANE_DEAL_ETAPY = [
    "Closed Won", "WdroÅ¼enie", "Trial", "WdroÅ¼eni Klienci"
]

# --- Funkcje pomocnicze z get_zoho_account_details.py (zminimalizowane do trybu kompaktowego) ---
def fmt_date_pl(value: Optional[str]) -> str:
    if not value:
        return ""
    s = str(value).strip()
    try:
        if s.endswith("Z"):
            s = s[:-1] + "+00:00"
        dt = datetime.fromisoformat(s)
        return dt.strftime("%d-%m-%Y")
    except Exception:
        try:
            dt = datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
            return dt.strftime("%d-%m-%Y")
        except Exception:
            if len(s) >= 10 and s[4] == '-' and s[7] == '-':
                y, m, d = s[:10].split('-')
                return f"{d}-{m}-{y}"
    return ""

def sanitize_for_csv_text(value: Any) -> str:
    """Usuwa znaki nowej linii dla CSV, zastÄ™pujÄ…c je separatorem ' | '."""
    try:
        s = "" if value is None else str(value)
        return s.replace("\r\n", " | ").replace("\n", " | ").replace("\r", " | ")
    except Exception:
        return ""

def _chunk_list(items: List[str], size: int = 50) -> List[List[str]]:
    return [items[i:i+size] for i in range(0, len(items), size)]

def _coql_escape(value: str) -> str:
    try:
        return str(value).replace("\\", "\\\\").replace("'", "\\'")
    except Exception:
        return ""

def _maybe_refresh_token_cooldown() -> Optional[str]:
    """OdÅ›wieÅ¼a token tylko jeÅ›li minÄ™Å‚o >= 15s od ostatniego odÅ›wieÅ¼enia.
    Zwraca nowy token albo None przy bÅ‚Ä™dzie.
    """
    try:
        last = float(ACCESS_TOKEN_HOLDER.get("refreshed_at") or 0.0)
        if time.time() - last < 15.0:
            return ACCESS_TOKEN_HOLDER.get("token")
        new_tok = get_access_token()
        return new_tok
    except Exception as e:
        logging.error(f"Cooldown refresh failed: {e}")
        return None

def execute_api_request(access_token: str, url: str) -> Optional[Dict[str, Any]]:
    def _do_request(tok: str) -> Optional[Dict[str, Any]]:
        headers = {"Authorization": f"Zoho-oauthtoken {tok}"}
        request = urllib.request.Request(url, headers=headers, method="GET")
        time.sleep(REQUEST_SLEEP_S)
        logging.info(f"HTTP GET request: {url}")
        with urllib.request.urlopen(request, timeout=30) as response:
            data = response.read().decode("utf-8")
            if not data:
                logging.info("HTTP GET response: <empty>")
                return None
            logging.info(f"HTTP GET response JSON: {data}")
            return json.loads(data)
    # pierwsza prÃ³ba
    try:
        # uÅ¼ywaj zawsze najÅ›wieÅ¼szego tokena z holdera
        tok0 = ACCESS_TOKEN_HOLDER.get("token") or access_token
        return _do_request(tok0)
    except urllib.error.HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            pass
        logging.error(f"HTTPError GET {e.code} for {url}: {body}")
        if e.code in (401, 403):
            # odÅ›wieÅ¼ token i ponÃ³w raz
            try:
                new_tok = _maybe_refresh_token_cooldown() or access_token
                return _do_request(new_tok)
            except Exception as e2:
                logging.error(f"Ponowienie po odÅ›wieÅ¼eniu tokena nie powiodÅ‚o siÄ™: {e2}")
                return None
        return None
    except Exception as e:
        # Retry na typowe timeouty socket/WinError 10060
        err_text = str(e)
        if "timed out" in err_text or "10060" in err_text:
            for i in range(MAX_RETRY_TIMEOUT):
                time.sleep(2 + i * 3)
                try:
                    tok = ACCESS_TOKEN_HOLDER.get("token") or access_token
                    return _do_request(tok)
                except Exception:
                    continue
        logging.error(f"BÅ‚Ä…d podczas wykonywania zapytania do API ({url}): {e}")
        return None

def execute_coql_query(access_token: str, query: str) -> Optional[List[Dict[str, Any]]]:
    def _do_post(tok: str) -> Optional[List[Dict[str, Any]]]:
        payload = json.dumps({"select_query": query}).encode("utf-8")
        headers = {
            "Authorization": f"Zoho-oauthtoken {tok}",
            "Content-Type": "application/json",
        }
        req = urllib.request.Request(ZOHO_CRM_COQL_URL, data=payload, headers=headers, method="POST")
        time.sleep(REQUEST_SLEEP_S)
        logging.info(f"COQL request JSON: {json.dumps({'select_query': query}, ensure_ascii=False)}")
        with urllib.request.urlopen(req, timeout=30) as resp:
            text = resp.read().decode("utf-8")
            if not text:
                return []
            logging.info(f"COQL response JSON: {text}")
            return json.loads(text).get("data", [])
    try:
        tok0 = ACCESS_TOKEN_HOLDER.get("token") or access_token
        return _do_post(tok0)
    except urllib.error.HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8", errors="replace")
        except Exception:
            pass
        logging.error(f"HTTPError COQL {e.code}: {body} | {query}")
        if e.code in (401, 403):
            try:
                new_tok = _maybe_refresh_token_cooldown() or access_token
                return _do_post(new_tok)
            except Exception as e2:
                logging.error(f"Ponowienie COQL po odÅ›wieÅ¼eniu tokena nie powiodÅ‚o siÄ™: {e2}")
                return None
        return None
    except Exception as e:
        # Retry na typowe timeouty socket/WinError 10060
        err_text = str(e)
        if "timed out" in err_text or "10060" in err_text:
            for i in range(MAX_RETRY_TIMEOUT):
                time.sleep(2 + i * 3)
                try:
                    tok = ACCESS_TOKEN_HOLDER.get("token") or access_token
                    return _do_post(tok)
                except Exception:
                    continue
        logging.error(f"BÅ‚Ä…d COQL: {e} | {query}")
        return None

def search_module_by_relation(
    access_token: str,
    module: str,
    field_api: str,
    related_id: str,
    fields: List[str],
) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    page = 1
    while True:
        criteria = f"({field_api}:equals:{related_id})"
        params = {
            "criteria": criteria,
            "fields": ",".join(fields),
            "per_page": 200,
            "page": page,
        }
        url = f"{ZOHO_CRM_API_BASE_URL}/{module}/search?{urllib.parse.urlencode(params)}"
        payload = execute_api_request(access_token, url)
        if not payload:
            break
        data = payload.get("data", [])
        records.extend(data)
        info = payload.get("info", {})
        if not info or not info.get("more_records"):
            break
        page += 1
    return records

def get_account_details_by_nip(access_token: str, nip: str) -> Optional[Dict[str, Any]]:
    logging.info(f"Wyszukiwanie kont (wszystkich) dla NIP: {nip}...")
    rows = execute_coql_query(access_token, (
        "select id, Status_klienta, Parent_Account, Firma_NIP, Nazwa_zwyczajowa, Account_Name from Accounts "
        f"where Firma_NIP = '{nip}'"
    )) or []
    if not rows:
        logging.warning("Nie znaleziono konta o podanym numerze NIP.")
        return None
    def has_parent(rec: Dict[str, Any]) -> bool:
        p = rec.get("Parent_Account")
        if isinstance(p, dict):
            return bool(p.get("id"))
        if isinstance(p, str):
            return bool(p)
        return False
    root_candidates = [r for r in rows if not has_parent(r)]
    chosen = root_candidates[0] if root_candidates else rows[0]
    logging.info(f"Wybrano konto gÅ‚Ã³wne: {chosen.get('id')} (spoÅ›rÃ³d {len(rows)} rekordÃ³w dla NIP)")
    return chosen

def get_account_name(access_token: str, account_id: str) -> str:
    rows = execute_coql_query(access_token, f"select Account_Name from Accounts where id = {account_id}") or []
    return str(rows[0].get("Account_Name") or "") if rows else ""

def get_display_name(access_token: str, main_account: Dict[str, Any]) -> str:
    custom = str(main_account.get("Nazwa_zwyczajowa") or "").strip()
    if custom:
        return custom
    parent_field = main_account.get("Parent_Account")
    parent_id = None
    if isinstance(parent_field, dict):
        parent_id = parent_field.get("id")
    elif isinstance(parent_field, str) and parent_field:
        parent_id = parent_field
    if parent_id:
        return get_account_name(access_token, parent_id) or ""
    return get_account_name(access_token, str(main_account.get("id") or "")) or ""

def find_related_account_ids(access_token: str, main_account: Dict[str, Any]) -> List[str]:
    """Zwraca peÅ‚ne drzewo kont: od korzenia (bez rodzica) zbiera wszystkie potomne,
    dodatkowo dokÅ‚ada wszystkie konta z tym samym NIPem.
    """
    def get_parent_id(acc_id: str) -> Optional[str]:
        rows = execute_coql_query(access_token, f"select Parent_Account from Accounts where id = {acc_id}") or []
        if not rows:
            return None
        p = rows[0].get("Parent_Account")
        if isinstance(p, dict) and p.get("id"):
            return str(p.get("id"))
        if isinstance(p, str) and p:
            return p
        return None

    main_id = str(main_account.get("id") or "")
    if not main_id:
        return []

    # znajdÅº korzeÅ„ (bez rodzica)
    root_id = main_id
    while True:
        pid = get_parent_id(root_id)
        if not pid:
            break
        root_id = pid

    # BFS po caÅ‚ym drzewie potomkÃ³w
    seen: set = set()
    result: List[str] = []
    queue: List[str] = [root_id]
    while queue:
        current = queue.pop(0)
        if current in seen:
            continue
        seen.add(current)
        result.append(current)
        rows = execute_coql_query(access_token, f"select id from Accounts where Parent_Account.id = {current}") or []
        for r in rows:
            cid = str(r.get("id")) if r.get("id") else None
            if cid and cid not in seen:
                queue.append(cid)

    # DoÅ‚Ã³Å¼ konta o tym samym NIP co konto gÅ‚Ã³wne
    nip_val = main_account.get("Firma_NIP")
    if nip_val:
        same = execute_coql_query(access_token, f"select id from Accounts where Firma_NIP = '{nip_val}'") or []
        for r in same:
            rid = str(r.get("id")) if r.get("id") else None
            if rid and rid not in seen:
                result.append(rid)
                seen.add(rid)

    logging.info(f"PowiÄ…zane konta (drzewo od korzenia {root_id}): {len(result)} -> {', '.join(result)}")
    return result

def get_related_records(access_token: str, module: str, record_id: str, fields: List[str]) -> List[Dict[str, Any]]:
    url = f"{ZOHO_CRM_API_BASE_URL}/Accounts/{record_id}/{module}?fields={','.join(fields)}&per_page=200"
    resp = execute_api_request(access_token, url)
    return resp.get("data", []) if resp else []

def search_module_by_relation(access_token: str, module: str, field_api: str, related_id: str, fields: List[str]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    page = 1
    while True:
        params = {
            "criteria": f"({field_api}:equals:{related_id})",
            "fields": ",".join(fields),
            "per_page": 200,
            "page": page,
        }
        url = f"{ZOHO_CRM_API_BASE_URL}/{module}/search?{urllib.parse.urlencode(params)}"
        payload = execute_api_request(access_token, url)
        if not payload:
            break
        records.extend(payload.get("data", []))
        info = payload.get("info", {})
        if not info or not info.get("more_records"):
            break
        page += 1
    return records

def get_all_activities(access_token: str, account_ids: List[str], contact_ids: List[str], lead_ids: List[str], deal_ids: List[str]) -> List[Dict[str, Any]]:
    def coql_fetch_single(module: str, field: str, rid: str, fields: List[str]) -> List[Dict[str, Any]]:
        # COQL fallback: tylko dla Who_Id (Leads/Contacts). What_Id daje 400 dla wielu moduÅ‚Ã³w.
        if field != "Who_Id":
            return []
        fields_sel = ", ".join(fields)
        query = f"select {fields_sel} from {module} where {field}.id in ({rid})"
        rows = execute_coql_query(access_token, query) or []
        return rows

    def fetch(module: str, field: str, ids: List[str]) -> List[Dict[str, Any]]:
        out: List[Dict[str, Any]] = []
        call_counter = 0
        for i in ids:
            call_counter += 1
            if call_counter % 20 == 0:
                time.sleep(0.8)
            # 1) prÃ³ba przez /{module}/search
            recs = search_module_by_relation(access_token, module, field, i, [
                "id",
                ("Event_Title" if module=="Events" else "Subject"),
                "Description",
                "Who_Id",
                "What_Id",
                ("Call_Start_Time" if module=="Calls" else ("Start_DateTime" if module=="Events" else "Due_Date"))
            ])
            # 2) fallback COQL, gdy search nic nie zwrÃ³ciÅ‚
            if not recs:
                recs = coql_fetch_single(
                    module,
                    field,
                    i,
                    [
                        "id",
                        ("Event_Title" if module=="Events" else "Subject"),
                        ("Call_Start_Time" if module=="Calls" else ("Start_DateTime" if module=="Events" else "Due_Date")),
                        "Description",
                        "Who_Id",
                        "What_Id",
                    ],
                )
            out += recs
        return out
    # Accounts / Contacts
    calls = fetch("Calls", "What_Id", account_ids) + fetch("Calls", "Who_Id", contact_ids)
    events = fetch("Events", "What_Id", account_ids) + fetch("Events", "Who_Id", contact_ids)
    tasks = fetch("Tasks", "What_Id", account_ids) + fetch("Tasks", "Who_Id", contact_ids)
    # Deals (What_Id) + fallback relacje /Deals/{id}/...
    calls += fetch("Calls", "What_Id", deal_ids)
    events += fetch("Events", "What_Id", deal_ids)
    tasks += fetch("Tasks", "What_Id", deal_ids)
    for did in deal_ids:
        time.sleep(0.1)
        for rel_mod, date_field in [("Calls", "Call_Start_Time"), ("Events", "Start_DateTime"), ("Tasks", "Due_Date")]:
            rel_fields = (["id", "Event_Title", date_field, "Description", "Who_Id", "What_Id"] if rel_mod == "Events" else ["id", "Subject", date_field, "Description", "Who_Id", "What_Id"])
            fields_str = ",".join(rel_fields)
            page = 1
            while True:
                rel_url = f"{ZOHO_CRM_API_BASE_URL}/Deals/{did}/{rel_mod}?fields={fields_str}&per_page=200&page={page}"
                payload = execute_api_request(access_token, rel_url)
                if not payload:
                    break
                data = payload.get("data", [])
                if rel_mod == "Calls":
                    calls += data
                elif rel_mod == "Events":
                    events += data
                else:
                    tasks += data
                info = payload.get("info", {})
                if not info or not info.get("more_records"):
                    break
                page += 1
    # Leads â€“ Who_Id (standard) + TEST: What_Id (eksperymentalnie, moÅ¼e zwrÃ³ciÄ‡ pusto) + fallback relacje /Leads/{id}/...
    calls += fetch("Calls", "Who_Id", lead_ids)
    events += fetch("Events", "Who_Id", lead_ids)
    tasks += fetch("Tasks", "Who_Id", lead_ids)
    # TEST What_Id dla LeadÃ³w
    calls += fetch("Calls", "What_Id", lead_ids)
    events += fetch("Events", "What_Id", lead_ids)
    tasks += fetch("Tasks", "What_Id", lead_ids)
    for lid in lead_ids:
        for rel_mod, date_field in [("Calls", "Call_Start_Time"), ("Events", "Start_DateTime"), ("Tasks", "Due_Date")]:
            rel_fields = (["Event_Title", date_field, "Description", "Who_Id", "What_Id"] if rel_mod == "Events" else ["Subject", date_field, "Description", "Who_Id", "What_Id"])
            fields_str = ",".join(rel_fields)
            page = 1
            while True:
                rel_url = f"{ZOHO_CRM_API_BASE_URL}/Leads/{lid}/{rel_mod}?fields={fields_str}&per_page=200&page={page}"
                payload = execute_api_request(access_token, rel_url)
                if not payload:
                    break
                data = payload.get("data", [])
                if rel_mod == "Calls":
                    calls += data
                elif rel_mod == "Events":
                    events += data
                else:
                    tasks += data
                info = payload.get("info", {})
                if not info or not info.get("more_records"):
                    break
                page += 1
    def norm(module: str, rec: Dict[str, Any]) -> Dict[str, Any]:
        if module == "Calls":
            when = rec.get("Call_Start_Time")
        elif module == "Events":
            when = rec.get("Start_DateTime")
        else:
            when = rec.get("Due_Date")
        symbol = "ğŸ§©"
        who = rec.get("Who_Id")
        what = rec.get("What_Id")
        if isinstance(who, dict):
            mod = (who.get("module") or "").lower()
            if mod == "contacts": symbol = "ğŸ‘¤"
            elif mod == "leads": symbol = "ğŸ§²"
        if isinstance(what, dict):
            mod = (what.get("module") or "").lower()
            if mod == "accounts": symbol = "ğŸ¢"
            elif mod == "deals": symbol = "ğŸ¤"
        subj_val = rec.get("Event_Title") if module == "Events" else rec.get("Subject", "")
        desc_val = rec.get("Description", "")
        if isinstance(desc_val, str) and len(desc_val) > 300:
            desc_val = desc_val[:180]
        return {"id": str(rec.get("id") or ""), "type": module[:-1] if module.endswith("s") else module, "time": when or "", "related_symbol": symbol, "Subject": subj_val or "", "Description": desc_val or ""}
    all_items: List[Dict[str, Any]] = [*map(lambda r: norm("Calls", r), calls), *map(lambda r: norm("Events", r), events), *map(lambda r: norm("Tasks", r), tasks)]
    all_items.sort(key=lambda x: x.get("time", ""), reverse=True)
    return all_items

def format_compact_output(details: Dict[str, Any]) -> str:
    parts: List[str] = []
    if details.get('display_name'):
        parts.append(f"Nazwa: {details['display_name']}")
    parts.append(f"ğŸ“„ {details.get('status_klienta','Brak')}")
    parts.append("-" * 20)
    line2: List[str] = []
    if int(details.get('contacts_count', 0) or 0):
        line2.append(f"ğŸ‘¥ {int(details.get('contacts_count', 0) or 0)}")
    if int(details.get('related_accounts_count', 0) or 0):
        line2.append(f"ğŸ¢ {int(details.get('related_accounts_count', 0) or 0)}")
    if line2:
        parts.append(" ".join(line2))
    if int(details.get('leads_count', 0) or 0):
        line = f"ğŸ§² {int(details.get('leads_count', 0) or 0)}"
        bracket_parts: List[str] = []
        disq = int((details.get('status_counts') or {}).get('leads_disqualified', 0))
        if disq:
            dt = (details.get('last_dates') or {}).get('lead_disqualified', '')
            bracket_parts.append(f"âŒ {disq}{(' ' + dt) if dt else ''}")
        active = int((details.get('status_counts') or {}).get('leads_active', 0))
        if active:
            bracket_parts.append(f"âš¡ {active}")
        if bracket_parts:
            line += " [" + " | ".join(bracket_parts) + "]"
        parts.append(line)
    if int(details.get('deals_count', 0) or 0):
        line = f"ğŸ¯ {int(details.get('deals_count', 0) or 0)}"
        bracket_parts: List[str] = []
        lost = int((details.get('status_counts') or {}).get('deals_lost', 0))
        if lost:
            lost_dt = (details.get('last_dates') or {}).get('deal_lost', '')
            bracket_parts.append(f"âŒ {lost}{(' ' + lost_dt) if lost_dt else ''}")
        active = int((details.get('status_counts') or {}).get('deals_active', 0))
        if active:
            bracket_parts.append(f"âš¡ {active}")
        won = int((details.get('status_counts') or {}).get('deals_won', 0))
        if won:
            won_dt = (details.get('last_dates') or {}).get('deal_won', '')
            bracket_parts.append(f"ğŸ† {won}{(' ' + won_dt) if won_dt else ''}")
        if bracket_parts:
            line += " [" + " | ".join(bracket_parts) + "]"
        parts.append(line)
    parts.append("-" * 20)
    all_activities: List[Dict[str, Any]] = details.get('all_activities') or []
    if details.get('full_activities'):
        for a in all_activities:
            a_type = a.get('type', 'N/A')
            type_icon = 'ğŸ§©' if a_type == 'Task' else ('ğŸ“' if a_type == 'Call' else ('ğŸ“…' if a_type == 'Event' else 'ğŸ§©'))
            date_pl = fmt_date_pl(a.get('time',''))
            subject = a.get('Subject', 'Brak tematu')
            parts.append(f"  - {date_pl} | {type_icon} -> {subject}")
    else:
        if all_activities:
            seen = set()
            task_count = 0
            call_count = 0
            event_count = 0
            max_tasks = 1
            max_calls = 5
            max_events = 4
            for a in all_activities:
                key = (a.get('type'), fmt_date_pl(a.get('time','')), (a.get('Subject') or '').strip())
                if key in seen:
                    continue
                a_type = a.get('type')
                if a_type == 'Task':
                    if task_count >= max_tasks:
                        continue
                    task_count += 1
                    type_icon = 'ğŸ§©'
                elif a_type == 'Call':
                    if call_count >= max_calls:
                        continue
                    call_count += 1
                    type_icon = 'ğŸ“'
                elif a_type == 'Event':
                    if event_count >= max_events:
                        continue
                    event_count += 1
                    type_icon = 'ğŸ“…'
                else:
                    type_icon = 'ğŸ§©'
                seen.add(key)
                date_pl = fmt_date_pl(a.get('time',''))
                subject = a.get('Subject','Brak tematu')
                parts.append(f"  - {date_pl} | {type_icon} -> {subject}")
        else:
            parts.append("Brak zarejestrowanych aktywnoÅ›ci.")
    return "\n".join(parts)

def write_detailed_reports_for_nip(access_token: str, nip: str, base_run_dir: str) -> None:
    main = get_account_details_by_nip(access_token, nip)
    if not main:
        return
    related_ids = find_related_account_ids(access_token, main)
    # katalog dla NIP: NIP_NAZWA (bezpieczna nazwa)
    display_name = get_display_name(access_token, main)
    safe_disp = "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in str(display_name) or "")
    safe_disp = safe_disp.strip("_")
    nip_folder_name = f"{nip}_{safe_disp}" if safe_disp else str(nip)
    nip_dir = os.path.join(base_run_dir, nip_folder_name)
    os.makedirs(nip_dir, exist_ok=True)
    # AGREGAT GRUPY
    contacts: List[Dict[str, Any]] = []
    deals: List[Dict[str, Any]] = []
    for aid in related_ids:
        contacts += get_related_records(access_token, "Contacts", aid, ["id"])
        deals += get_related_records(access_token, "Deals", aid, ["id", "Stage", "Modified_Time"])
    ids_in = ", ".join(related_ids)
    leads = execute_coql_query(access_token, f"select id, Lead_Status, Modified_Time from Leads where Firma_w_bazie.id in ({ids_in})") or []
    contact_ids = [c['id'] for c in contacts if c.get('id')]
    deal_ids = [d['id'] for d in deals if d.get('id')]
    activities = get_all_activities(access_token, related_ids, contact_ids, [l.get('id') for l in leads if l.get('id')], deal_ids)
    def latest_date(items: List[Dict[str, Any]], pred) -> str:
        dates = [fmt_date_pl(i.get("Modified_Time")) for i in items if pred(i) and i.get("Modified_Time")]
        return max(dates) if dates else ""
    status_klienta = str(main.get("Status_klienta") or "Brak")
    group_details: Dict[str, Any] = {
        "display_name": display_name,
        "status_klienta": status_klienta,
        "contacts_count": len(contacts),
        "related_accounts_count": len(related_ids),
        "leads_count": len(leads),
        "deals_count": len(deals),
        "all_activities": activities,
        "status_counts": {
            "leads_active": sum(1 for l in leads if l.get("Lead_Status") in AKTYWNE_LEAD_STATUSY),
            "leads_disqualified": sum(1 for l in leads if l.get("Lead_Status") == "Zdyskwalifikowany"),
            "deals_active": sum(1 for d in deals if d.get("Stage") in AKTYWNE_DEAL_ETAPY),
            "deals_won": sum(1 for d in deals if d.get("Stage") in WYGRANE_DEAL_ETAPY),
            "deals_lost": sum(1 for d in deals if d.get("Stage") == "Closed Lost"),
        },
        "last_dates": {
            "lead_disqualified": latest_date(leads, lambda l: l.get("Lead_Status") == "Zdyskwalifikowany"),
            "deal_lost": latest_date(deals, lambda d: d.get("Stage") == "Closed Lost"),
            "deal_won": latest_date(deals, lambda d: d.get("Stage") in WYGRANE_DEAL_ETAPY),
        },
    }
    raport_spojny_text = format_compact_output(group_details)
    with open(os.path.join(nip_dir, "raport_spojny.txt"), "w", encoding="utf-8-sig") as f:
        f.write("### RAPORT ZBIORCZY â€“ AGREGAT GRUPY\n\n")
        f.write(raport_spojny_text)
        f.write("\n\n### RAPORTY PER-PLACÃ“WKA\n\n")
    # PER-PLACÃ“WKA
    for acc_id in related_ids:
        acc_name = get_account_name(access_token, acc_id) or acc_id
        acc_contacts = get_related_records(access_token, "Contacts", acc_id, ["id"])
        acc_deals = get_related_records(access_token, "Deals", acc_id, ["id", "Stage", "Modified_Time"])
        acc_leads = execute_coql_query(access_token, f"select id, Lead_Status, Modified_Time from Leads where Firma_w_bazie.id = {acc_id}") or []
        acc_contact_ids = [c.get('id') for c in acc_contacts if c.get('id')]
        acc_deal_ids = [d.get('id') for d in acc_deals if d.get('id')]
        acc_acts = get_all_activities(access_token, [acc_id], acc_contact_ids, [l.get('id') for l in acc_leads if l.get('id')], acc_deal_ids)
        def latest_date2(items: List[Dict[str, Any]], pred) -> str:
            dates = [fmt_date_pl(i.get("Modified_Time")) for i in items if pred(i) and i.get("Modified_Time")]
            return max(dates) if dates else ""
        acc_details: Dict[str, Any] = {
            "display_name": acc_name,
            "status_klienta": str(get_account_details_by_nip(access_token, nip).get("Status_klienta") or "Brak"),
            "contacts_count": len(acc_contacts),
            "related_accounts_count": 1,
            "leads_count": len(acc_leads),
            "deals_count": len(acc_deals),
            "all_activities": acc_acts,
            "full_activities": True,
            "status_counts": {
                "leads_active": sum(1 for l in acc_leads if l.get("Lead_Status") in AKTYWNE_LEAD_STATUSY),
                "leads_disqualified": sum(1 for l in acc_leads if l.get("Lead_Status") == "Zdyskwalifikowany"),
                "deals_active": sum(1 for d in acc_deals if d.get("Stage") in AKTYWNE_DEAL_ETAPY),
                "deals_won": sum(1 for d in acc_deals if d.get("Stage") in WYGRANE_DEAL_ETAPY),
                "deals_lost": sum(1 for d in acc_deals if d.get("Stage") == "Closed Lost"),
            },
            "last_dates": {
                "lead_disqualified": latest_date2(acc_leads, lambda l: l.get("Lead_Status") == "Zdyskwalifikowany"),
                "deal_lost": latest_date2(acc_deals, lambda d: d.get("Stage") == "Closed Lost"),
                "deal_won": latest_date2(acc_deals, lambda d: d.get("Stage") in WYGRANE_DEAL_ETAPY),
            },
        }
        acc_text = format_compact_output(acc_details)
        safe_name = "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in str(acc_name) or str(acc_id))
        acc_file = os.path.join(nip_dir, f"wynik_{safe_name}.txt")
        with open(acc_file, "w", encoding="utf-8-sig") as f:
            f.write(f"PlacÃ³wka: {acc_name} (ID: {acc_id})\n")
            f.write(acc_text)
        # dopisz do zbiorczego
        with open(os.path.join(nip_dir, "raport_spojny.txt"), "a", encoding="utf-8-sig") as f:
            f.write(f"\n--- {acc_name} (ID: {acc_id}) ---\n\n")
            f.write(f"PlacÃ³wka: {acc_name} (ID: {acc_id})\n")
            f.write(acc_text)

def build_compact_summary(access_token: str, nip: str) -> Tuple[str, Dict[str, Any]]:
    main = get_account_details_by_nip(access_token, nip)
    if not main:
        return "", {"last_call_event_date": None, "last_call_event_date_fmt": "", "greenlight_text": "NIE - brak rekordu", "has_active_lead": False, "has_active_deal": False, "is_customer": False}
    related_ids = find_related_account_ids(access_token, main)
    contacts: List[Dict[str, Any]] = []
    deals: List[Dict[str, Any]] = []
    for aid in related_ids:
        contacts += get_related_records(access_token, "Contacts", aid, ["id"])
        deals += get_related_records(access_token, "Deals", aid, ["id", "Stage", "Modified_Time"])
    # leads po lookup id konta (batching <=50 ID) + fallback po nazwie firmy
    leads: List[Dict[str, Any]] = []
    seen_leads: set = set()
    for batch in _chunk_list(related_ids, 50):
        ids_in = ", ".join(batch)
        rows = execute_coql_query(access_token, f"select id, Lead_Status, Modified_Time from Leads where Firma_w_bazie.id in ({ids_in})") or []
        for r in rows:
            lid = r.get("id")
            if lid and lid not in seen_leads:
                seen_leads.add(lid)
                leads.append(r)
    # Uwaga: Å›wiadomie WYÅÄ„CZONE wyszukiwanie LeadÃ³w po nazwie firmy (tylko po ID lookup)
    contact_ids = [c['id'] for c in contacts if c.get('id')]
    deal_ids = [d['id'] for d in deals if d.get('id')]
    activities = get_all_activities(access_token, related_ids, contact_ids, [l.get('id') for l in leads if l.get('id')], deal_ids)
    # zliczenia i daty
    def latest_date(items: List[Dict[str, Any]], pred) -> str:
        dates = [fmt_date_pl(i.get("Modified_Time")) for i in items if pred(i) and i.get("Modified_Time")]
        return max(dates) if dates else ""
    # Status klienta â€“ zbiorczo dla caÅ‚ej grupy (jeÅ›li gdziekolwiek 'jest' â†’ 'jest')
    base_status = str(main.get("Status_klienta") or "Brak")
    group_status = base_status
    try:
        for batch in _chunk_list(related_ids, 50):
            ids_in = ", ".join(batch)
            srows = execute_coql_query(access_token, f"select Status_klienta from Accounts where id in ({ids_in})") or []
            statuses = [str(r.get("Status_klienta") or "") for r in srows]
            if any(s == "jest" for s in statuses):
                group_status = "jest"
                break
            # jeÅ›li brak 'jest', zachowujemy pierwszy niepusty jako reprezentatywny
            for s in statuses:
                if s:
                    group_status = s
                    break
    except Exception:
        pass
    display_name = get_display_name(access_token, main)
    details: Dict[str, Any] = {
        "display_name": display_name,
        "status_klienta": group_status,
        "contacts_count": len(contacts),
        "related_accounts_count": len(related_ids),
        "leads_count": len(leads),
        "deals_count": len(deals),
        "all_activities": activities,
        "status_counts": {
            "leads_active": sum(1 for l in leads if l.get("Lead_Status") in AKTYWNE_LEAD_STATUSY),
            "leads_disqualified": sum(1 for l in leads if l.get("Lead_Status") == "Zdyskwalifikowany"),
            "deals_active": sum(1 for d in deals if d.get("Stage") in AKTYWNE_DEAL_ETAPY),
            "deals_won": sum(1 for d in deals if d.get("Stage") in WYGRANE_DEAL_ETAPY),
            "deals_lost": sum(1 for d in deals if d.get("Stage") == "Closed Lost"),
        },
        "last_dates": {
            "lead_disqualified": latest_date(leads, lambda l: l.get("Lead_Status") == "Zdyskwalifikowany"),
            "deal_lost": latest_date(deals, lambda d: d.get("Stage") == "Closed Lost"),
            "deal_won": latest_date(deals, lambda d: d.get("Stage") in WYGRANE_DEAL_ETAPY),
        },
    }
    # dodatkowe: ostatnia aktywnoÅ›Ä‡ Call/Spotkanie â€“ tylko z PRZESZÅOÅšCI (ignoruj przyszÅ‚e zaplanowane)
    last_call_event_iso: Optional[str] = None
    now_ref = datetime.now()
    for a in activities:
        if a.get('type') in ('Call', 'Event') and a.get('time'):
            cand = a.get('time')
            dt_cand = None
            # uÅ¼ywamy tego samego parsera co dalej
            try:
                t = cand.strip()
                if t.endswith("Z"):
                    t = t[:-1] + "+00:00"
                dt_cand = datetime.fromisoformat(t)
            except Exception:
                try:
                    dt_cand = datetime.strptime(cand[:19], "%Y-%m-%d %H:%M:%S")
                except Exception:
                    dt_cand = None
            if dt_cand is None:
                continue
            # jeÅ›li bez strefy â€“ porÃ³wnujemy do lokalnego now
            if dt_cand.tzinfo:
                if datetime.now(dt_cand.tzinfo) >= dt_cand:
                    last_call_event_iso = cand
                    break
            else:
                if now_ref >= dt_cand:
                    last_call_event_iso = cand
                    break
    last_call_event_fmt = fmt_date_pl(last_call_event_iso) if last_call_event_iso else ""

    # Zielone Å›wiatÅ‚o dla Migtel
    # warunek 1: ostatni call/meeting > 80 dni temu (brak traktujemy jako speÅ‚niony)
    def parse_dt_iso(s: Optional[str]) -> Optional[datetime]:
        if not s:
            return None
        try:
            t = s.strip()
            if t.endswith("Z"):
                t = t[:-1] + "+00:00"
            return datetime.fromisoformat(t)
        except Exception:
            try:
                return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
            except Exception:
                return None
    dt_last = parse_dt_iso(last_call_event_iso)
    days_ok = True
    if dt_last is not None:
        diff_days = (datetime.now(dt_last.tzinfo) - dt_last).days if dt_last.tzinfo else (datetime.now() - dt_last).days
        days_ok = diff_days > 80
    # warunek 2: nie jest klientem â€“ WYÅÄ„CZNIE po Status_klienta (nie po wygranym Dealu)
    deals_won_count = details['status_counts'].get('deals_won', 0)
    has_won_deal = deals_won_count > 0
    is_customer = (group_status == 'jest')
    not_customer_ok = not is_customer
    # warunek 3 i 4: brak aktywnych leadÃ³w/deali
    has_active_lead = details['status_counts'].get('leads_active', 0) > 0 if 'leads_active' in details['status_counts'] else any(l.get("Lead_Status") in AKTYWNE_LEAD_STATUSY for l in leads)
    has_active_deal = details['status_counts'].get('deals_active', 0) > 0 if 'deals_active' in details['status_counts'] else any(d.get("Stage") in AKTYWNE_DEAL_ETAPY for d in deals)
    no_active_lead_ok = not has_active_lead
    no_active_deal_ok = not has_active_deal
    reasons: List[str] = []
    if not days_ok:
        reasons.append("ostatnia aktywnoÅ›Ä‡ <= 80 dni")
    if not not_customer_ok:
        reasons.append("klient 'jest'")
    if not no_active_lead_ok:
        reasons.append("aktywny Lead")
    if not no_active_deal_ok:
        reasons.append("aktywny Deal")
    greenlight_text = "TAK" if (days_ok and not_customer_ok and no_active_lead_ok and no_active_deal_ok) else ("NIE - " + ", ".join(reasons) if reasons else "NIE")

    # format kompaktowy â€“ bez nagÅ‚Ã³wkÃ³w
    parts: List[str] = []
    if details.get('display_name'):
        parts.append(f"Nazwa: {details['display_name']}")
    parts.append(f"ğŸ“„ {details.get('status_klienta','Brak')}")
    parts.append("-" * 20)
    line2 = []
    if details['contacts_count']:
        line2.append(f"ğŸ‘¥ {details['contacts_count']}")
    if details['related_accounts_count']:
        line2.append(f"ğŸ¢ {details['related_accounts_count']}")
    if line2:
        parts.append(" ".join(line2))
    if details['leads_count']:
        line = f"ğŸ§² {details['leads_count']}"
        bracket_parts: List[str] = []
        disq = details['status_counts'].get('leads_disqualified', 0)
        if disq:
            dt = details['last_dates'].get('lead_disqualified', '')
            bracket_parts.append(f"âŒ {disq}{(' ' + dt) if dt else ''}")
        active = details['status_counts'].get('leads_active', 0)
        if active:
            bracket_parts.append(f"âš¡ {active}")
        if bracket_parts:
            line += " [" + " | ".join(bracket_parts) + "]"
        parts.append(line)
    if details['deals_count']:
        line = f"ğŸ¯ {details['deals_count']}"
        bracket_parts: List[str] = []
        lost = details['status_counts'].get('deals_lost', 0)
        if lost:
            lost_dt = details['last_dates'].get('deal_lost', '')
            bracket_parts.append(f"âŒ {lost}{(' ' + lost_dt) if lost_dt else ''}")
        active = details['status_counts'].get('deals_active', 0)
        if active:
            bracket_parts.append(f"âš¡ {active}")
        won = details['status_counts'].get('deals_won', 0)
        if won:
            won_dt = details['last_dates'].get('deal_won', '')
            bracket_parts.append(f"ğŸ† {won}{(' ' + won_dt) if won_dt else ''}")
        if bracket_parts:
            line += " [" + " | ".join(bracket_parts) + "]"
        parts.append(line)
    parts.append("-" * 20)
    # Historia: zbiorczo â€“ filtr: bez pustych opisÃ³w, max 2 taski, deduplikacja
    if details['all_activities']:
        seen = set()
        task_count = 0
        call_count = 0
        event_count = 0
        max_tasks = 1
        max_calls = 5
        max_events = 4
        for a in details['all_activities']:
            key = (a.get('type'), fmt_date_pl(a.get('time','')), (a.get('Subject') or '').strip())
            if key in seen:
                continue
            a_type = a.get('type')
            if a_type == 'Task':
                if task_count >= max_tasks:
                    continue
                task_count += 1
                type_icon = 'ğŸ§©'
            elif a_type == 'Call':
                if call_count >= max_calls:
                    continue
                call_count += 1
                type_icon = 'ğŸ“'
            elif a_type == 'Event':
                if event_count >= max_events:
                    continue
                event_count += 1
                type_icon = 'ğŸ“…'
            else:
                type_icon = 'ğŸ§©'
            seen.add(key)
            date_pl = fmt_date_pl(a.get('time',''))
            subject = a.get('Subject','Brak tematu')
            parts.append(f"  - {date_pl} | {type_icon} -> {subject}")
    else:
        parts.append("Brak zarejestrowanych aktywnoÅ›ci.")
    doubt_flag = has_won_deal and (group_status != 'jest')
    return "\n".join(parts), {
        "last_call_event_date": last_call_event_iso,
        "last_call_event_date_fmt": last_call_event_fmt,
        "greenlight_text": greenlight_text,
        "has_active_lead": has_active_lead,
        "has_active_deal": has_active_deal,
        "is_customer": is_customer,
        "group_status": group_status,
        "has_won_deal": has_won_deal,
        "doubt": doubt_flag,
    }


ZOHO_CRM_COQL_URL = "https://www.zohoapis.eu/crm/v3/coql"


class TqdmLoggingHandler(logging.Handler):
    def __init__(self, level=logging.NOTSET):
        super().__init__(level)

    def emit(self, record):
        try:
            msg = self.format(record)
            tqdm.write(msg, file=sys.stderr)
            self.flush()
        except (KeyboardInterrupt, SystemExit):
            raise
        except Exception:
            self.handleError(record)


def setup_logging(log_dir: str) -> None:
    """Konfiguruje system logowania do pliku i konsoli z obsÅ‚ugÄ… tqdm."""
    log_file = os.path.join(log_dir, "log.txt")
    
    # Usuwamy domyÅ›lne handlery, aby uniknÄ…Ä‡ podwÃ³jnego logowania
    root_logger = logging.getLogger()
    if root_logger.hasHandlers():
        root_logger.handlers.clear()

    handlers = [logging.FileHandler(log_file, encoding="utf-8")]
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] - %(message)s",
        handlers=handlers,
    )

# Kontrola wypisywania statusÃ³w perâ€‘NIP przez logger do konsoli
# (pozostawiamy False, a linie wypisujemy rÄ™cznie obok paska)
SHOW_PER_NIP_CONSOLE = False

def write_console_status_dual(message_emoji: str, message_ascii: str) -> None:
    """Druk linii nad paskiem z prÃ³bÄ… emoji i bezpiecznym fallbackiem ASCII."""
    try:
        tqdm.write(message_emoji, file=sys.stderr)
    except Exception:
        try:
            tqdm.write(message_ascii, file=sys.stderr)
        except Exception:
            pass


def get_access_token() -> str:
    """
    Pobiera token dostÄ™powy Zoho, korzystajÄ…c z konfiguracji w zmiennych Å›rodowiskowych.
    """
    logging.info("Pobieranie tokena dostÄ™powego Zoho...")
    client_id = os.environ.get("ZOHO_MEDIDESK_CLIENT_ID", "WPISZ_CLIENT_ID")
    client_secret = os.environ.get("ZOHO_MEDIDESK_CLIENT_SECRET", "WPISZ_CLIENT_SECRET")
    refresh_token = os.environ.get("ZOHO_MEDIDESK_REFRESH_TOKEN", "WPISZ_REFRESH_TOKEN")

    if any(val.startswith("WPISZ_") for val in [client_id, client_secret, refresh_token]):
        logging.error(
            "BÅ‚Ä…d: Dane (client_id, client_secret, refresh_token) nie zostaÅ‚y skonfigurowane."
        )
        logging.error(
            "Ustaw zmienne Å›rodowiskowe (np. ZOHO_MEDIDESK_CLIENT_ID) lub wpisz je bezpoÅ›rednio w pliku."
        )
        sys.exit(1)

    try:
        token_info = refresh_access_token(
            client_id=client_id,
            client_secret=client_secret,
            refresh_token=refresh_token,
        )
        logging.info("PomyÅ›lnie pobrano token.")
        ACCESS_TOKEN_HOLDER["token"] = token_info["access_token"]
        ACCESS_TOKEN_HOLDER["refreshed_at"] = time.time()
        return token_info["access_token"]
    except RuntimeError as e:
        logging.error(f"WystÄ…piÅ‚ bÅ‚Ä…d podczas pobierania tokena: {e}")
        sys.exit(1)


def is_valid_nip(nip: str) -> bool:
    """
    Sprawdza poprawnoÅ›Ä‡ polskiego numeru NIP (dÅ‚ugoÅ›Ä‡, cyfry, suma kontrolna).
    """
    if pd.isna(nip):
        return False
    nip = str(nip).strip().replace("-", "")
    if len(nip) != 10 or not nip.isdigit():
        return False

    weights = (6, 5, 7, 2, 3, 4, 5, 6, 7)
    nip_digits = [int(d) for d in nip]
    checksum = sum(w * d for w, d in zip(weights, nip_digits[:-1]))
    return (checksum % 11) == nip_digits[-1]


def verify_nip_in_zoho(access_token: str, nip: str, retries: int = 2, delay: int = 2) -> Tuple[str, Optional[str]]:
    """
    Sprawdza, czy w Zoho CRM istnieje konto z podanym NIP-em, z mechanizmem ponawiania.

    Args:
        access_token: Token dostÄ™powy do API Zoho.
        nip: Numer NIP do sprawdzenia.
        retries: Liczba prÃ³b ponowienia w przypadku bÅ‚Ä™du.
        delay: Czas oczekiwania (w sekundach) przed ponowieniem.

    Returns:
        Krotka (status, record_id), np. ("ISTNIEJE", "123..."), ("NIE ISTNIEJE", None), ("BÅÄ„D", None).
    """
    if pd.isna(nip) or not str(nip).strip():
        return "BRAK NIP", None
    
    clean_nip = str(nip).strip()

    query = f"select id from Accounts where Firma_NIP = '{clean_nip}'"
    data = {"select_query": query}
    encoded_data = json.dumps(data).encode("utf-8")

    headers = {
        "Authorization": f"Zoho-oauthtoken {access_token}",
        "Content-Type": "application/json",
    }
    def build_request(tok: str) -> urllib.request.Request:
        return urllib.request.Request(
            ZOHO_CRM_COQL_URL, data=encoded_data, headers={
                "Authorization": f"Zoho-oauthtoken {tok}",
                "Content-Type": "application/json",
            }, method="POST"
        )
    
    for attempt in range(retries + 1):
        try:
            tok0 = ACCESS_TOKEN_HOLDER.get("token") or access_token
            with urllib.request.urlopen(build_request(tok0), timeout=30) as response:
                payload_text = response.read().decode("utf-8")

                # ZMIANA LOGIKI: Pusta odpowiedÅº jest interpretowana jako "NIE ISTNIEJE"
                if not payload_text:
                    return "NIE ISTNIEJE", None
                
                # JeÅ›li odpowiedÅº nie jest pusta, prÃ³bujemy jÄ… zinterpretowaÄ‡ jako JSON
                payload = json.loads(payload_text)
                if "data" in payload and len(payload["data"]) > 0:
                    return "ISTNIEJE", payload["data"][0]["id"]
                
                # JeÅ›li jest JSON, ale bez klucza 'data' lub 'data' jest puste
                return "NIE ISTNIEJE", None
                
        except urllib.error.HTTPError as http_err:
            error_body = http_err.read().decode("utf-8", errors="replace")
            logging.error(
                f"BÅ‚Ä…d API Zoho dla NIP {clean_nip} (prÃ³ba {attempt + 1}/{retries + 1}): {http_err.code} - {error_body}"
            )
            # Przy bÅ‚Ä™dach autoryzacji (401, 403) nie ma sensu ponawiaÄ‡
            if http_err.code in [401, 403]:
                try:
                    new_tok = _maybe_refresh_token_cooldown() or access_token
                    with urllib.request.urlopen(build_request(new_tok), timeout=30) as response:
                        payload_text = response.read().decode("utf-8")
                        if not payload_text:
                            return "NIE ISTNIEJE", None
                        payload = json.loads(payload_text)
                        if "data" in payload and len(payload["data"]) > 0:
                            return "ISTNIEJE", payload["data"][0]["id"]
                        return "NIE ISTNIEJE", None
                except Exception as e2:
                    logging.error(f"Ponowienie verify po odÅ›wieÅ¼eniu tokena nie powiodÅ‚o siÄ™: {e2}")
                    return "BÅÄ„D", None
            if attempt < retries:
                logging.info(f"Ponawiam prÃ³bÄ™ za {delay} sek...")
                time.sleep(delay)
        except (urllib.error.URLError, TimeoutError) as err:
            # Prawdziwe bÅ‚Ä™dy poÅ‚Ä…czenia - ponawiamy
            logging.error(f"BÅ‚Ä…d poÅ‚Ä…czenia dla NIP {clean_nip} (prÃ³ba {attempt + 1}/{retries + 1}): {err}")
            if attempt < retries:
                logging.info(f"Ponawiam prÃ³bÄ™ za {delay} sek...")
                time.sleep(delay)
        except json.JSONDecodeError as err:
            # BÅ‚Ä…d parsowania JSON - to jest teraz prawdziwy bÅ‚Ä…d, bo odpowiedÅº nie byÅ‚a pusta
            logging.error(f"BÅ‚Ä…d odpowiedzi JSON dla NIP {clean_nip} (prÃ³ba {attempt + 1}/{retries + 1}): {err}")
            if attempt < retries:
                logging.info(f"Ponawiam prÃ³bÄ™ za {delay} sek...")
                time.sleep(delay)


    return "BÅÄ„D", None


def main() -> None:
    """
    GÅ‚Ã³wna funkcja skryptu.
    """
    parser = argparse.ArgumentParser(
        description="Weryfikuje NIP-y z pliku CSV/XLSX w Zoho CRM i zapisuje wyniki do dedykowanego katalogu."
    )
    parser.add_argument(
        "input_file",
        nargs="?",
        default=None,
        help="ÅšcieÅ¼ka do pliku wejÅ›ciowego. JeÅ›li pominiÄ™ta, program poprosi o niÄ….",
    )
    args = parser.parse_args()

    input_file = args.input_file
    if not input_file:
        raw_path = input(
            "ProszÄ™ przeciÄ…gnÄ…Ä‡ plik do okna konsoli lub wkleiÄ‡ Å›cieÅ¼kÄ™ i nacisnÄ…Ä‡ Enter: "
        )
        # Zaawansowane czyszczenie Å›cieÅ¼ki z potencjalnych artefaktÃ³w z konsoli (np. PowerShell)
        clean_path = raw_path.strip()
        if clean_path.startswith("& "):
            clean_path = clean_path[2:].strip()
        input_file = clean_path.strip("'\"")


    if not os.path.exists(input_file):
        print(f"BÅ‚Ä…d: Plik '{input_file}' nie istnieje. ProszÄ™ sprawdziÄ‡ Å›cieÅ¼kÄ™ i sprÃ³bowaÄ‡ ponownie.", file=sys.stderr)
        sys.exit(1)

    # 1. Tworzenie katalogÃ³w
    base_output_dir = "wyniki_czy_istnieje"
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    run_dir = os.path.join(base_output_dir, timestamp)
    os.makedirs(run_dir, exist_ok=True)

    # 2. Konfiguracja logowania
    setup_logging(run_dir)

    # 3. Logika aplikacji
    logging.info(f"Rozpoczynam weryfikacjÄ™ NIP-Ã³w z pliku: {input_file}")
    logging.info(f"Wyniki zostanÄ… zapisane w katalogu: {run_dir}")

    try:
        # Wczytywanie pliku wejÅ›ciowego (CSV lub Excel)
        file_ext = os.path.splitext(input_file)[1].lower()
        if file_ext == '.csv':
            df = pd.read_csv(input_file, dtype=str, sep=';')
        elif file_ext in ['.xlsx', '.xls']:
            df = pd.read_excel(input_file, dtype=str)
        else:
            logging.error(f"NieobsÅ‚ugiwany format pliku: {file_ext}. UÅ¼yj .csv lub .xlsx.")
            sys.exit(1)

        # Sprawdzenie, czy plik jest pusty
        if df.empty:
            logging.warning("Plik wejÅ›ciowy jest pusty.")
            return
            
        nip_column_name = df.columns[0]
        logging.info(f"Odczytano NIP-y z kolumny: '{nip_column_name}'")

        access_token = get_access_token()
        # zapisz takÅ¼e do holdera dla automatycznego odÅ›wieÅ¼ania
        ACCESS_TOKEN_HOLDER["token"] = access_token
        
        # Czekamy 2 sekundy, aby daÄ‡ czas serwerom Zoho na peÅ‚nÄ… aktywacjÄ™ nowego tokena
        logging.info("Oczekiwanie 2s na peÅ‚nÄ… aktywacjÄ™ tokena...")
        time.sleep(2)

        results = []
        compact_summaries: List[str] = []
        last_call_event_dates: List[str] = []
        greenlight_values: List[str] = []
        group_statuses: List[str] = []
        has_won_deal_flags: List[str] = []
        doubts: List[str] = []
        validation_results = []
        summary_counts = {"ISTNIEJE": 0, "NIE ISTNIEJE": 0, "NIEPRAWIDÅOWY": 0, "BÅÄ„D": 0}

        # ÅšcieÅ¼ki wynikÃ³w i helper do autosave CSV (co 50)
        output_base_path = os.path.join(run_dir, "wynik")
        csv_path = f"{output_base_path}.csv"
        xlsx_path = f"{output_base_path}.xlsx"
        xlsx_partial_path = f"{output_base_path}_partial.xlsx"

        def save_partial_csv(processed_count: int) -> None:
            try:
                part = df.iloc[:processed_count].copy()
                statuses = [r.get("status", "") for r in results[:processed_count]]
                ids = [(r.get("record_id") or "") for r in results[:processed_count]]
                part.insert(1, "Podsumowanie 360", compact_summaries[:processed_count])
                part.insert(2, "Ostatnia aktywnoÅ›Ä‡ Call/Spotkanie", last_call_event_dates[:processed_count])
                part.insert(3, "Zielone Å›wiatÅ‚o dla Migtel", greenlight_values[:processed_count])
                # (opcjonalnie, jeÅ›li zebrane) kolumny grupowe
                try:
                    part.insert(4, "Status klienta (grupa)", group_statuses[:processed_count])
                    part.insert(5, "Wygrany Deal (grupa)", has_won_deal_flags[:processed_count])
                    part.insert(6, "WÄ…tpliwoÅ›Ä‡ (Deal wygrany, a status != 'jest')", doubts[:processed_count])
                    insert_offset = 3
                except Exception:
                    insert_offset = 0
                part.insert(4 + insert_offset, "Walidacja NIP", validation_results[:processed_count])
                part.insert(5 + insert_offset, "Status weryfikacji ZOHO", statuses)
                part.insert(6 + insert_offset, "ID rekordu ZOHO", ids)
                # CSV autosave: bez nowych linii w komÃ³rkach
                if "Podsumowanie 360" in part.columns:
                    part["Podsumowanie 360"] = part["Podsumowanie 360"].map(sanitize_for_csv_text)
                part.to_csv(csv_path, index=False, encoding="utf-8-sig")
                logging.info(f"Autozapis CSV po {processed_count} wierszach: {csv_path}")
            except Exception as e:
                logging.error(f"BÅ‚Ä…d autosave CSV: {e}")

        def save_partial_xlsx(processed_count: int) -> None:
            try:
                part = df.iloc[:processed_count].copy()
                statuses = [r.get("status", "") for r in results[:processed_count]]
                ids = [(r.get("record_id") or "") for r in results[:processed_count]]
                part.insert(1, "Podsumowanie 360", compact_summaries[:processed_count])
                part.insert(2, "Ostatnia aktywnoÅ›Ä‡ Call/Spotkanie", last_call_event_dates[:processed_count])
                part.insert(3, "Zielone Å›wiatÅ‚o dla Migtel", greenlight_values[:processed_count])
                try:
                    part.insert(4, "Status klienta (grupa)", group_statuses[:processed_count])
                    part.insert(5, "Wygrany Deal (grupa)", has_won_deal_flags[:processed_count])
                    part.insert(6, "WÄ…tpliwoÅ›Ä‡ (Deal wygrany, a status != 'jest')", doubts[:processed_count])
                    insert_offset = 3
                except Exception:
                    insert_offset = 0
                part.insert(4 + insert_offset, "Walidacja NIP", validation_results[:processed_count])
                part.insert(5 + insert_offset, "Status weryfikacji ZOHO", statuses)
                part.insert(6 + insert_offset, "ID rekordu ZOHO", ids)
                part.to_excel(xlsx_partial_path, index=False, engine='openpyxl')
                # wrap text dla kolumny Podsumowanie 360 (B)
                try:
                    wb = load_workbook(xlsx_partial_path)
                    ws = wb.active
                    for cell in ws['B']:
                        if cell.row == 1:
                            continue
                        cell.alignment = Alignment(wrap_text=True)
                    wb.save(xlsx_partial_path)
                except Exception:
                    pass
                logging.info(f"Autozapis XLSX po {processed_count} wierszach: {xlsx_partial_path}")
            except Exception as e:
                logging.error(f"BÅ‚Ä…d autosave XLSX: {e}")
        # UÅ¼ycie tqdm do stworzenia paska postÄ™pu
        with tqdm(total=len(df), desc="Weryfikacja NIP-Ã³w", unit=" wiersz") as pbar:
            for index, row in df.iterrows():
                nip_value = row[nip_column_name]
                logging.info(f"START NIP {nip_value}")
                
                # Krok 1: Walidacja poprawnoÅ›ci NIP
                if is_valid_nip(nip_value):
                    validation_status = "PRAWIDÅOWY"
                    # Krok 2: Sprawdzanie w Zoho (tylko jeÅ›li NIP jest poprawny)
                    status, record_id = verify_nip_in_zoho(access_token, nip_value)
                else:
                    validation_status = "NIEPRAWIDÅOWY"
                    status, record_id = "POMINIÄ˜TO", None

                # Zliczanie wynikÃ³w do podsumowania
                if status == "ISTNIEJE":
                    summary_counts["ISTNIEJE"] += 1
                elif status == "NIE ISTNIEJE":
                    summary_counts["NIE ISTNIEJE"] += 1
                elif status == "POMINIÄ˜TO":
                    summary_counts["NIEPRAWIDÅOWY"] += 1
                elif status == "BÅÄ„D":
                    summary_counts["BÅÄ„D"] += 1

                # Podsumowanie 360 â€“ tylko gdy istnieje rekord
                compact_text = ""
                if validation_status == "PRAWIDÅOWY" and status == "ISTNIEJE":
                    try:
                        compact_text, extra = build_compact_summary(access_token, str(nip_value).strip())
                        last_call_event_dates.append(extra.get("last_call_event_date_fmt", ""))
                        greenlight_values.append(extra.get("greenlight_text", "NIE"))
                        group_statuses.append(extra.get("group_status", ""))
                        has_won_deal_flags.append("TAK" if extra.get("has_won_deal") else "NIE")
                        doubts.append("TAK" if extra.get("doubt") else "NIE")
                    except Exception as e:
                        logging.error(f"BÅ‚Ä…d budowania podsumowania 360 dla NIP {nip_value}: {e}")
                        compact_text = ""
                        last_call_event_dates.append("")
                        greenlight_values.append("NIE - bÅ‚Ä…d")
                        group_statuses.append("")
                        has_won_deal_flags.append("NIE")
                        doubts.append("NIE")
                elif status == "NIE ISTNIEJE":
                    # Nie ma w Zoho -> Zielone Å›wiatÅ‚o = TAK; brak daty aktywnoÅ›ci
                    compact_text = ""
                    last_call_event_dates.append("")
                    greenlight_values.append("TAK")
                    group_statuses.append("")
                    has_won_deal_flags.append("NIE")
                    doubts.append("NIE")
                # jeÅ›li istnieje â€“ zapisz teÅ¼ szczegÃ³Å‚y per NIP do podfolderu
                try:
                    if status == "ISTNIEJE":
                        write_detailed_reports_for_nip(access_token, str(nip_value).strip(), run_dir)
                except Exception as e:
                    logging.error(f"BÅ‚Ä…d zapisu raportÃ³w szczegÃ³Å‚owych dla NIP {nip_value}: {e}")

                compact_summaries.append(compact_text)
                if len(last_call_event_dates) < len(compact_summaries):
                    last_call_event_dates.append("")
                if len(greenlight_values) < len(compact_summaries):
                    greenlight_values.append("NIE")
                if len(group_statuses) < len(compact_summaries):
                    group_statuses.append("")
                if len(has_won_deal_flags) < len(compact_summaries):
                    has_won_deal_flags.append("NIE")
                if len(doubts) < len(compact_summaries):
                    doubts.append("NIE")

                validation_results.append(validation_status)
                results.append({"status": status, "record_id": record_id})
                
                # Log (z emoji) i konsola ASCII
                if status == "ISTNIEJE":
                    log_status = f"âœ… NIP {nip_value} -> ISTNIEJE"
                    console_emoji = f"âœ… NIP {nip_value} -> ISTNIEJE"
                    console_ascii = f"[OK] NIP {nip_value} -> ISTNIEJE"
                elif status == "NIE ISTNIEJE":
                    log_status = f"âŒ NIP {nip_value} -> NIE ISTNIEJE"
                    console_emoji = f"âŒ NIP {nip_value} -> NIE ISTNIEJE"
                    console_ascii = f"[NIE] NIP {nip_value} -> NIE ISTNIEJE"
                elif status == "POMINIÄ˜TO":
                    log_status = f"â¡ï¸  NIP {nip_value} -> {validation_status} (PominiÄ™to Zoho)"
                    console_emoji = f"â¡ NIP {nip_value} -> {validation_status} (PominiÄ™to)"
                    console_ascii = f"[POMINIÄ˜TO] NIP {nip_value} -> {validation_status}"
                else:
                    log_status = f"âš ï¸  NIP {nip_value} -> {status}"
                    console_emoji = f"âš  NIP {nip_value} -> {status}"
                    console_ascii = f"[UWAGA] NIP {nip_value} -> {status}"

                logging.info(log_status)
                write_console_status_dual(console_emoji, console_ascii)
                pbar.update(1)
                logging.info(f"END NIP {nip_value}")
                
                # Dodajemy pauzÄ™, aby uniknÄ…Ä‡ przekroczenia limitÃ³w API Zoho
                time.sleep(0.4)

                # Autosave co 50 rekordÃ³w (CSV kumulacyjnie)
                processed = index + 1
                if processed % 50 == 0:
                    save_partial_csv(processed)
                    save_partial_xlsx(processed)

        # Dodawanie nowych kolumn
        logging.info("Tworzenie plikÃ³w wynikowych...")
        results_df = pd.DataFrame(results)
        # Nowe kolumny â€“ przed innymi wstawianymi kolumnami
        df.insert(1, "Podsumowanie 360", compact_summaries)
        df.insert(2, "Ostatnia aktywnoÅ›Ä‡ Call/Spotkanie", last_call_event_dates)
        df.insert(3, "Zielone Å›wiatÅ‚o dla Migtel", greenlight_values)
        df.insert(4, "Status klienta (grupa)", group_statuses)
        df.insert(5, "Wygrany Deal (grupa)", has_won_deal_flags)
        df.insert(6, "WÄ…tpliwoÅ›Ä‡ (Deal wygrany, a status != 'jest')", doubts)
        df.insert(7, "Walidacja NIP", validation_results)
        df.insert(8, "Status weryfikacji ZOHO", results_df["status"])
        df.insert(9, "ID rekordu ZOHO", results_df["record_id"].fillna(""))
        
        # Zapisywanie wynikÃ³w (final)

        # CSV: bez nowych linii w komÃ³rkach (np. w Podsumowaniu 360)
        df_csv = df.copy()
        if "Podsumowanie 360" in df_csv.columns:
            df_csv["Podsumowanie 360"] = df_csv["Podsumowanie 360"].map(sanitize_for_csv_text)
        df_csv.to_csv(csv_path, index=False, encoding="utf-8-sig")
        logging.info(f"Zapisano wyniki do pliku CSV: {csv_path}")
        
        df.to_excel(xlsx_path, index=False, engine='openpyxl')
        logging.info(f"Zapisano wyniki do pliku XLSX: {xlsx_path}")

        # WÅ‚Ä…cz zawijanie tekstu (wrap text), aby \n byÅ‚y widoczne jako nowe linie w Excelu
        try:
            wb = load_workbook(xlsx_path)
            ws = wb.active
            # Kolumna B to "Podsumowanie 360" (dodawana jako insert(1, ...))
            for cell in ws['B']:
                if cell.row == 1:
                    continue
                cell.alignment = Alignment(wrap_text=True)
            wb.save(xlsx_path)
            logging.info("Ustawiono zawijanie tekstu w kolumnie 'Podsumowanie 360' w XLSX.")
        except Exception as e:
            logging.warning(f"Nie udaÅ‚o siÄ™ ustawiÄ‡ zawijania tekstu w XLSX: {e}")

        # --- PODSUMOWANIE ---
        total_rows = len(df)
        summary_log_message = (
            f"Podsumowanie: Przetworzono {total_rows} wierszy. "
            f"IstniejÄ…ce: {summary_counts['ISTNIEJE']}, "
            f"NieistniejÄ…ce: {summary_counts['NIE ISTNIEJE']}, "
            f"NieprawidÅ‚owe: {summary_counts['NIEPRAWIDÅOWY']}, "
            f"BÅ‚Ä™dy: {summary_counts['BÅÄ„D']}."
        )
        logging.info(summary_log_message)

        # Przygotowanie danych do Å‚adnej tabeli w konsoli
        summary_data = {
            "Przetworzono wierszy": total_rows,
            "IstniejÄ…ce w Zoho": summary_counts['ISTNIEJE'],
            "NieistniejÄ…ce w Zoho": summary_counts['NIE ISTNIEJE'],
            "NieprawidÅ‚owe NIP-y": summary_counts['NIEPRAWIDÅOWY'],
            "BÅ‚Ä™dy weryfikacji": summary_counts['BÅÄ„D']
        }

        # Prosty ASCII-only komunikat (bez ramek/emoji) â€“ bez wypisywania przez tqdm, by uniknÄ…Ä‡ problemÃ³w kodowania
        summary_console_message = (
            f"\nPODSUMOWANIE: Przetworzono {total_rows} | "
            f"IstniejÄ…ce: {summary_counts['ISTNIEJE']} | "
            f"NieistniejÄ…ce: {summary_counts['NIE ISTNIEJE']} | "
            f"NieprawidÅ‚owe: {summary_counts['NIEPRAWIDÅOWY']} | "
            f"BÅ‚Ä™dy: {summary_counts['BÅÄ„D']}\n"
        )
        logging.info(summary_console_message)

        logging.info("ZakoÅ„czono pomyÅ›lnie.")

    except FileNotFoundError:
        logging.error(f"BÅ‚Ä…d: Plik wejÅ›ciowy '{input_file}' nie zostaÅ‚ znaleziony.")
        sys.exit(1)
    except Exception as e:
        logging.error(f"WystÄ…piÅ‚ nieoczekiwany bÅ‚Ä…d: {e}", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()

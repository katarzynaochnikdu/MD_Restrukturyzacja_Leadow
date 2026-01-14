"""
Buduje XLSX z wygenerowanych tabel funnel.

Wejścia (CSV w curated_data/):
- funnel_entry_month_source.csv
- funnel_entry_month_global.csv
- funnel_source_global.csv

Wyjście:
- Funnel_entry_2025.xlsx (albo wersjonowany jeśli plik jest otwarty)
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import CURATED_DATA_DIR, ANALYSIS_YEAR


IN_MONTH_SOURCE = Path(CURATED_DATA_DIR) / "funnel_entry_month_source.csv"
IN_MONTH_GLOBAL = Path(CURATED_DATA_DIR) / "funnel_entry_month_global.csv"
IN_SOURCE_GLOBAL = Path(CURATED_DATA_DIR) / "funnel_source_global.csv"


POLISH_COLS: Dict[str, str] = {
    "project_type": "Typ projektu (PreSales/Sales only)",
    "entry_month": "Miesiąc wejścia (wg Created_Time)",
    "source_I_tier": "Źródło – poziom I",
    "source_II_tier": "Źródło – poziom II",
    "source_III_tier": "Źródło – poziom III",
    "source_IV_tier": "Źródło – poziom IV",
    "N_entered": "Liczba projektów",
    "N_with_meeting": "Liczba Meeting I",
    "pct_with_meeting": "Odsetek Meeting I",
    "median_days_to_meeting": "Mediana dni do 1. spotkania",
    "mean_days_to_meeting": "Średnia dni do 1. spotkania",
    "N_to_deal": "Liczba Deali z projektów",
    "pct_to_deal": "Odsetek projektów z Dealam",
    "N_won": "Liczba umów",
    "pct_won_of_entered": "Odsetek umów z wejść",
    "win_rate_of_deals": "Win rate: umowy / deale",
    "avg_ojp": "Średnia ocena OJP (Ankiety)",
    "sum_ankiety": "Suma ankiet",
    "avg_ojp_weighted": "Średnia ocena OJP ważona ankietami",
}


def _safe_output_path(base: Path) -> Path:
    if not base.exists():
        return base
    for i in range(2, 100):
        candidate = base.with_name(f"{base.stem}_v{i}{base.suffix}")
        if not candidate.exists():
            return candidate
    return base.with_name(f"{base.stem}_vX{base.suffix}")


def _autosize_columns(ws, max_width: int = 60) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, min(ws.max_row + 1, 3000)):  # cap for speed
            v = ws.cell(row, col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), max_width)


def _write_table(ws, df: pd.DataFrame, header_fill: PatternFill | None = None) -> None:
    header_font = Font(bold=True)
    header_alignment = Alignment(vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=False)

    # Header
    for j, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col_name)
        cell.font = header_font
        cell.alignment = header_alignment
        if header_fill:
            cell.fill = header_fill

    # Body
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.alignment = body_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _apply_number_formats(ws, col_formats: Dict[str, str]) -> None:
    # Map header -> column idx
    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for header, fmt in col_formats.items():
        col = header_to_col.get(header)
        if not col:
            continue
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = fmt


def _group_month_source_sheet(ws, df: pd.DataFrame) -> None:
    """
    Robi grupowanie wierszy:
    - poziom 1: entry_month
    - poziom 2: source_I_tier + source_II_tier
    - poziom 3: source_III_tier (wiersze danych)
    """
    # Zakładamy że df jest posortowany: month, I, II, III
    # W Excelu grupujemy poprzez outlineLevel.
    # Będziemy dodawać „wiersze nagłówkowe” dla grup, a dane pod nimi.

    # Najpierw czyścimy sheet i budujemy od nowa w układzie grupowym:
    ws.delete_rows(1, ws.max_row)

    cols = [
        "project_type",
        "entry_month",
        "source_I_tier",
        "source_II_tier",
        "source_III_tier",
        "N_entered",
        "N_with_meeting",
        "pct_with_meeting",
        "median_days_to_meeting",
        "mean_days_to_meeting",
        "N_to_deal",
        "pct_to_deal",
        "N_won",
        "pct_won_of_entered",
        "win_rate_of_deals",
    ]
    df2 = df.copy()
    for c in cols:
        if c not in df2.columns:
            df2[c] = ""
    df2 = df2[cols]

    # Header (polskie, pełne opisy)
    header_fill = PatternFill("solid", fgColor="E7EEF9")
    header_font = Font(bold=True)
    header_alignment = Alignment(vertical="center", wrap_text=True)
    for j, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=POLISH_COLS.get(col_name, col_name))
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    current_row = 2

    def write_group_row(label: str, outline_level: int) -> int:
        nonlocal current_row
        ws.cell(current_row, 1, label)
        ws.row_dimensions[current_row].outlineLevel = outline_level
        ws.row_dimensions[current_row].collapsed = False
        # styl nagłówka grupy
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(current_row, c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
        current_row += 1
        return current_row - 1

    def write_data_row(values: List) -> int:
        nonlocal current_row
        for j, v in enumerate(values, start=1):
            ws.cell(current_row, j, v)
        ws.row_dimensions[current_row].outlineLevel = 4
        current_row += 1
        return current_row - 1

    # Grupowanie: project_type → entry_month → source
    for ptype, ptype_df in df2.groupby("project_type", dropna=False):
        ptype_label = str(ptype) if pd.notna(ptype) and str(ptype).strip() else "brak danych"
        ptype_row = write_group_row(f"Typ: {ptype_label}", outline_level=1)

        for month, month_df in ptype_df.groupby("entry_month", dropna=False):
            month_label = str(month) if pd.notna(month) and str(month).strip() else "brak miesiąca"
            month_row = write_group_row(f"  Miesiąc: {month_label}", outline_level=2)

            for (t1, t2), sub_df in month_df.groupby(["source_I_tier", "source_II_tier"], dropna=False):
                t1s = str(t1) if pd.notna(t1) and str(t1).strip() else "brak danych"
                t2s = str(t2) if pd.notna(t2) and str(t2).strip() else "brak danych"
                group_row = write_group_row(f"    Źródło (I/II): {t1s} / {t2s}", outline_level=3)

                for r in sub_df.itertuples(index=False):
                    write_data_row(list(r))

                ws.row_dimensions[group_row].collapsed = True

            ws.row_dimensions[month_row].collapsed = True

        ws.row_dimensions[ptype_row].collapsed = True

    # Ustawienia outline
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_properties.outlinePr.summaryRight = True


def main() -> None:
    for f in (IN_MONTH_SOURCE, IN_MONTH_GLOBAL, IN_SOURCE_GLOBAL):
        if not f.exists():
            raise FileNotFoundError(f"Brak {f}. Uruchom build_entry_funnel_by_source.py")

    df_month_source = pd.read_csv(IN_MONTH_SOURCE)
    df_month_global = pd.read_csv(IN_MONTH_GLOBAL)
    df_source_global = pd.read_csv(IN_SOURCE_GLOBAL)

    # Sortowanie dla grup
    df_month_source = df_month_source.sort_values(
        ["project_type", "entry_month", "source_I_tier", "source_II_tier", "source_III_tier"]
    )

    wb = Workbook()
    # Zakładka z opisami (żeby plik był czytelny „dla biznesu”)
    ws_info = wb.active
    ws_info.title = "Opis"
    ws_info["A1"] = f"Lejek wejścia {ANALYSIS_YEAR} – definicje i uwagi"
    ws_info["A1"].font = Font(bold=True, size=14)
    ws_info["A3"] = "Wejście (cykl):"
    ws_info["A4"] = "- zdeduplikowane ML/Lead (jeśli Lead pochodzi z ML) liczymy jako 1 projekt."
    ws_info["A6"] = "Miesiąc wejścia:"
    ws_info["A7"] = "- przypisany wg Created_Time rekordu wejściowego (ML albo Lead)."
    ws_info["A9"] = "Spotkanie:"
    ws_info["A10"] = "- status TAK, jeśli istnieje jakikolwiek rekord w related Events lub Events_History."
    ws_info["A12"] = "Dni do 1. spotkania:"
    ws_info["A13"] = "- różnica między momentem wejścia a czasem pierwszego spotkania (na etapie Leada)."
    ws_info["A15"] = "Deal / Umowa:"
    ws_info["A16"] = "- Deal: projekt ma Deal; Umowa: Deal w etapie umowy (Closed Won, Wdrożenie, Trial, Wdrożeni Klienci)."
    ws_info.column_dimensions["A"].width = 110

    ws_main = wb.create_sheet("Lejek_mies_zrodlo")

    # Główna zakładka: grupowanie wierszy
    _group_month_source_sheet(ws_main, df_month_source)

    # Format liczbowy
    _apply_number_formats(
        ws_main,
        {
            # Uwaga: w CSV procenty są w skali 0-100, więc używamy formatu "doklej %" (bez mnożenia x100).
            POLISH_COLS["pct_with_meeting"]: '0.00"%"',
            POLISH_COLS["pct_to_deal"]: '0.00"%"',
            POLISH_COLS["pct_won_of_entered"]: '0.00"%"',
            POLISH_COLS["win_rate_of_deals"]: '0.00"%"',
            POLISH_COLS["median_days_to_meeting"]: "0.00",
            POLISH_COLS["mean_days_to_meeting"]: "0.00",
        },
    )
    _autosize_columns(ws_main)

    # Zakładka global per miesiąc
    ws_m = wb.create_sheet("Lejek_mies_global")
    df_m = df_month_global.rename(columns=POLISH_COLS)
    _write_table(ws_m, df_m, header_fill=PatternFill("solid", fgColor="E7EEF9"))
    _apply_number_formats(
        ws_m,
        {
            POLISH_COLS["pct_with_meeting"]: '0.00"%"',
            POLISH_COLS["pct_to_deal"]: '0.00"%"',
            POLISH_COLS["pct_won_of_entered"]: '0.00"%"',
            POLISH_COLS["win_rate_of_deals"]: '0.00"%"',
            POLISH_COLS["median_days_to_meeting"]: "0.00",
            POLISH_COLS["mean_days_to_meeting"]: "0.00",
        },
    )
    _autosize_columns(ws_m)

    # Zakładka global per źródło + rozbicie miesięczne
    ws_s = wb.create_sheet("Lejek_zrodlo_global")
    
    # Dodaj 12 kolumn miesięcy (Styczeń-Grudzień) z liczbą projektów per miesiąc
    # Pivot: źródła (I/II/III) w wierszach, miesiące w kolumnach, wartość = N_entered
    months_order = ["2025-01", "2025-02", "2025-03", "2025-04", "2025-05", "2025-06",
                    "2025-07", "2025-08", "2025-09", "2025-10", "2025-11", "2025-12"]
    month_names = ["Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec",
                   "Lipiec", "Sierpień", "Wrzesień", "Październik", "Listopad", "Grudzień"]
    
    pivot_months = df_month_source.pivot_table(
        index=["project_type", "source_I_tier", "source_II_tier", "source_III_tier"],
        columns="entry_month",
        values="N_entered",
        aggfunc="sum",
        fill_value=0
    ).reset_index()
    
    # Zmień nazwy kolumn miesięcy na polskie
    pivot_months.columns = [
        c if c in ["project_type", "source_I_tier", "source_II_tier", "source_III_tier"]
        else month_names[months_order.index(c)] if c in months_order else c
        for c in pivot_months.columns
    ]
    
    # Merge z df_source_global (globalny funnel) po kluczach źródłowych
    df_s_with_months = df_source_global.merge(
        pivot_months,
        on=["project_type", "source_I_tier", "source_II_tier", "source_III_tier"],
        how="left"
    )
    
    # Zmień kolejność kolumn: project_type + źródła → miesiące → metryki globalne
    source_cols = ["project_type", "source_I_tier", "source_II_tier", "source_III_tier"]
    month_cols = [m for m in month_names if m in df_s_with_months.columns]
    metric_cols = [c for c in df_s_with_months.columns if c not in source_cols + month_cols]
    df_s_with_months = df_s_with_months[source_cols + month_cols + metric_cols]
    
    # Zmień nazwy kolumn źródłowych na polskie
    df_s = df_s_with_months.rename(columns=POLISH_COLS)
    _write_table(ws_s, df_s, header_fill=PatternFill("solid", fgColor="E7EEF9"))
    # Formatowanie liczb: miesiące jako liczby całkowite, %, dni jako decimal
    formats = {
        POLISH_COLS["pct_with_meeting"]: '0.00"%"',
        POLISH_COLS["pct_to_deal"]: '0.00"%"',
        POLISH_COLS["pct_won_of_entered"]: '0.00"%"',
        POLISH_COLS["win_rate_of_deals"]: '0.00"%"',
        POLISH_COLS["median_days_to_meeting"]: "0.00",
        POLISH_COLS["mean_days_to_meeting"]: "0.00",
        POLISH_COLS["avg_ojp"]: "0.00",
        POLISH_COLS["avg_ojp_weighted"]: "0.00",
        POLISH_COLS["sum_ankiety"]: "0",
    }
    # Dodaj format dla kolumn miesięcy (liczby całkowite)
    for month_name in month_names:
        formats[month_name] = "0"
    
    _apply_number_formats(ws_s, formats)
    _autosize_columns(ws_s)

    out = _safe_output_path(Path(f"Funnel_entry_{ANALYSIS_YEAR}.xlsx"))
    wb.save(out)
    print(f"Zapisano XLSX: {out}")


if __name__ == "__main__":
    main()


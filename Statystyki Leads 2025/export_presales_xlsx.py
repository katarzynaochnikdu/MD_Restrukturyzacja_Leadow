"""Export presales matching to XLSX with Zoho links."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Zoho base URLs
ZOHO_BASE = "https://crm.zoho.eu/crm/org20101283812/tab"

# Wczytaj dane - deal_id i account_id jako string
df = pd.read_csv('raw_data/presales_deals_matching.csv', dtype={'deal_id': str, 'account_id': str})

# Utwórz workbook
wb = Workbook()
ws = wb.active
ws.title = "Dopasowanie PreSales"

# Nagłówki
headers = [
    'Lp', 'Klient (lista PreSales)', 'Produkt', 'PreSales', 
    'Deal ID', 'Link do Deala', 'Account ID', 'Link do Account',
    'Nazwa firmy (Zoho)', 'Mój typ projektu', 
    'Owner Deala', 'Data utworzenia', 'Stage', 'Kwota',
    'Jakość dopasowania', 'Uwagi'
]

# Style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
link_font = Font(color="0563C1", underline="single")
warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Zapisz nagłówki
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Zapisz dane
for row_idx, row in df.iterrows():
    r = row_idx + 2
    
    deal_id = str(row['deal_id']) if pd.notna(row['deal_id']) and row['deal_id'] != '' else ''
    account_id = str(row['account_id']) if pd.notna(row['account_id']) and row['account_id'] != '' else ''
    
    # Dane podstawowe
    ws.cell(row=r, column=1, value=row['lp'])
    ws.cell(row=r, column=2, value=row['ps_klient'])
    ws.cell(row=r, column=3, value=row['ps_produkt'])
    ws.cell(row=r, column=4, value=row['ps_presales'])
    
    # Deal ID i link
    ws.cell(row=r, column=5, value=deal_id if deal_id else '')
    if deal_id:
        deal_url = f"{ZOHO_BASE}/Potentials/{deal_id}"
        cell = ws.cell(row=r, column=6, value=deal_url)
        cell.hyperlink = deal_url
        cell.font = link_font
    
    # Account ID i link
    ws.cell(row=r, column=7, value=account_id if account_id else '')
    if account_id and account_id != 'nan':
        acc_url = f"{ZOHO_BASE}/Accounts/{account_id}"
        cell = ws.cell(row=r, column=8, value=acc_url)
        cell.hyperlink = acc_url
        cell.font = link_font
    
    # Pozostałe dane
    ws.cell(row=r, column=9, value=row['account_name_zoho'] if pd.notna(row['account_name_zoho']) else '')
    ws.cell(row=r, column=10, value=row['moj_project_type'] if pd.notna(row['moj_project_type']) else '')
    ws.cell(row=r, column=11, value=row['deal_owner'] if pd.notna(row['deal_owner']) else '')
    ws.cell(row=r, column=12, value=row['deal_created'] if pd.notna(row['deal_created']) else '')
    ws.cell(row=r, column=13, value=row['deal_stage'] if pd.notna(row['deal_stage']) else '')
    ws.cell(row=r, column=14, value=row['deal_amount'] if pd.notna(row['deal_amount']) else '')
    ws.cell(row=r, column=15, value=row['match_quality'] if pd.notna(row['match_quality']) else '')
    ws.cell(row=r, column=16, value=row['uwagi'] if pd.notna(row['uwagi']) else '')
    
    # Kolorowanie wierszy
    uwagi = str(row['uwagi']) if pd.notna(row['uwagi']) else ''
    match_q = str(row['match_quality']) if pd.notna(row['match_quality']) else ''
    
    if 'RÓŻNICA' in uwagi or 'SPRAWDŹ' in uwagi:
        for col in range(1, 17):
            ws.cell(row=r, column=col).fill = warning_fill
    elif match_q == 'NIE ZNALEZIONO':
        for col in range(1, 17):
            ws.cell(row=r, column=col).fill = warning_fill
    elif 'DOKŁADNE' in match_q or '90%' in match_q:
        for col in range(1, 17):
            ws.cell(row=r, column=col).fill = ok_fill

# Szerokość kolumn
col_widths = [5, 30, 15, 20, 20, 50, 20, 50, 45, 15, 20, 12, 15, 10, 18, 35]
for col, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

# Zamroź nagłówek
ws.freeze_panes = 'A2'

# Zapisz
output_file = 'raw_data/presales_deals_matching.xlsx'
wb.save(output_file)
print(f"Zapisano: {output_file}")
print(f"Wierszy: {len(df)}")

"""
Scal pełną listę PreSales z danymi pomocniczymi i znajdź Deale.

LOGIKA:
1. Baza = sprzedaz_presales_source.csv (84 pozycje - pełna lista)
2. Pomocniczy = sprzedaz_presales_source_sztyczen_wrzesien_umowy.csv (58 z linkami)
3. Dla każdej pozycji z bazy:
   - Jeśli jest w pomocniczym → użyj linku do znalezienia Deala
   - Jeśli NIE → szukaj po nazwie w deals
"""

import pandas as pd
import json
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher

ZOHO_BASE = "https://crm.zoho.eu/crm/org20101283812/tab"

def extract_id_from_link(link):
    """Wyciągnij ID i typ z linku Zoho."""
    if not link or pd.isna(link):
        return None, None
    link = str(link).strip()
    patterns = [
        (r'/tab/Potentials/(\d+)', 'Deal'),
        (r'/tab/Leads/(\d+)', 'Lead'),
        (r'/tab/Accounts/(\d+)', 'Account'),
        (r'/tab/Events/(\d+)', 'Event'),
    ]
    for pattern, obj_type in patterns:
        match = re.search(pattern, link)
        if match:
            return match.group(1), obj_type
    return None, None

def normalize(s):
    """Normalizuj nazwę do porównania."""
    if not s:
        return ''
    s = str(s).lower()
    s = re.sub(r'[^a-ząćęłńóśźż0-9\s]', ' ', s)
    s = ' '.join(s.split())
    return s

def get_key_words(s):
    """Wyciągnij kluczowe słowa (>2 znaki)."""
    norm = normalize(s)
    return set(w for w in norm.split() if len(w) > 2)

def similarity(a, b):
    """Oblicz podobieństwo."""
    norm_a = normalize(a).replace(' ', '')
    norm_b = normalize(b).replace(' ', '')
    
    if norm_a == norm_b:
        return 1.0
    if norm_a in norm_b or norm_b in norm_a:
        return 0.95
    
    words_a = get_key_words(a)
    words_b = get_key_words(b)
    if words_a and words_b:
        common = words_a & words_b
        if len(common) >= 2:
            return 0.9
        if len(common) == 1 and len(words_a) <= 2:
            return 0.85
    
    return SequenceMatcher(None, norm_a, norm_b).ratio()

# ========== WCZYTAJ DANE ==========

# 1. Pełna lista PreSales (BAZA)
base_data = []
with open('raw_data/sprzedaz_presales_source.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)  # skip header
    for row in reader:
        if row and row[0].strip():
            base_data.append({
                'klient': row[0].strip(),
                'produkt': row[1].strip() if len(row) > 1 else '',
                'presales': row[2].strip() if len(row) > 2 else ''
            })

print(f"Pełna lista PreSales (baza): {len(base_data)} pozycji")

# 2. Dane pomocnicze z linkami (styczeń-wrzesień)
helper_data = {}
with open('raw_data/sprzedaz_presales_source_sztyczen_wrzesien_umowy.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)
    for row in reader:
        if row and len(row) > 1 and row[1].strip():
            klient_key = normalize(row[1].strip())
            link = row[7].strip() if len(row) > 7 else ''
            record_id, record_type = extract_id_from_link(link)
            helper_data[klient_key] = {
                'miesiac': row[0].strip(),
                'link': link,
                'record_id': record_id,
                'record_type': record_type,
                'typ_zrodla': row[8].strip() if len(row) > 8 else '',
                'zrodlo_szczegol': row[9].strip() if len(row) > 9 else '',
            }

print(f"Dane pomocnicze (sty-wrz): {len(helper_data)} pozycji z linkami")

# 3. Deals z lifecycle
lifecycle = pd.read_csv('curated_data/lifecycle_2025.csv', dtype=str)
deals_df = lifecycle[lifecycle['deal_id'].notna()].copy()

# 4. Raw deals
deals_raw = json.load(open('raw_data/deals_raw.json', 'r', encoding='utf-8'))
deals_map = {str(d.get('id')): d for d in deals_raw}

# Mapa account_name -> deals
account_to_deals = {}
for _, row in deals_df.iterrows():
    acc_name = row.get('account_name', '')
    acc_norm = normalize(acc_name)
    if acc_norm:
        if acc_norm not in account_to_deals:
            account_to_deals[acc_norm] = []
        account_to_deals[acc_norm].append({
            'deal_id': row['deal_id'],
            'account_id': row.get('account_id', ''),
            'account_name': acc_name,
            'deal_stage': row.get('deal_stage', ''),
            'deal_owner': row.get('deal_owner', ''),
            'project_type': row.get('project_type', '')
        })

print(f"Unikalne firmy w deals: {len(account_to_deals)}")

# ========== PRZETWARZANIE ==========

def find_deal_by_name(klient):
    """Szukaj Deala po nazwie firmy."""
    best_match = None
    best_score = 0
    best_acc_name = ''
    
    for acc_norm, deals_list in account_to_deals.items():
        orig_name = deals_list[0]['account_name'] if deals_list else ''
        score = similarity(klient, orig_name)
        if score > best_score:
            best_score = score
            best_match = deals_list
            best_acc_name = orig_name
    
    if best_match and best_score >= 0.5:
        # Preferuj wygrane deale
        won_stages = ['Closed Won', 'Wdrożenie', 'Trial', 'Wdrożeni Klienci']
        won_deals = [d for d in best_match if d['deal_stage'] in won_stages]
        deal_info = won_deals[0] if won_deals else best_match[0]
        return deal_info, best_score, best_acc_name
    
    return None, best_score, best_acc_name

results = []
for idx, base in enumerate(base_data, 1):
    klient = base['klient']
    klient_norm = normalize(klient)
    
    result = {
        'lp': idx,
        'klient': klient,
        'produkt': base['produkt'],
        'presales': base['presales'],
        'miesiac': '',
        'typ_zrodla': '',
        'zrodlo_szczegol': '',
        'link_zrodlowy': '',
        'link_type_zrodlowy': '',
        'deal_id': '',
        'account_id': '',
        'account_name_zoho': '',
        'deal_stage': '',
        'deal_owner': '',
        'project_type_moj': '',
        'link_deal': '',
        'link_account': '',
        'match_method': '',
        'uwagi': ''
    }
    
    # Sprawdź czy jest w danych pomocniczych
    helper = helper_data.get(klient_norm)
    
    if helper and helper['link']:
        result['miesiac'] = helper['miesiac']
        result['typ_zrodla'] = helper['typ_zrodla']
        result['zrodlo_szczegol'] = helper['zrodlo_szczegol']
        result['link_zrodlowy'] = helper['link']
        result['link_type_zrodlowy'] = helper['record_type']
        
        # Jeśli link to Deal - użyj bezpośrednio
        if helper['record_type'] == 'Deal' and helper['record_id']:
            deal = deals_map.get(helper['record_id'])
            if deal:
                result['deal_id'] = helper['record_id']
                acc = deal.get('Account_Name', {})
                result['account_id'] = str(acc.get('id', '')) if isinstance(acc, dict) else ''
                result['account_name_zoho'] = acc.get('name', '') if isinstance(acc, dict) else str(acc)
                result['deal_stage'] = deal.get('Stage', '')
                owner = deal.get('Owner', {})
                result['deal_owner'] = owner.get('name', '') if isinstance(owner, dict) else str(owner)
                result['link_deal'] = f"{ZOHO_BASE}/Potentials/{helper['record_id']}"
                if result['account_id']:
                    result['link_account'] = f"{ZOHO_BASE}/Accounts/{result['account_id']}"
                result['match_method'] = 'LINK->DEAL (bezpośredni)'
                
                lc_match = deals_df[deals_df['deal_id'] == helper['record_id']]
                if len(lc_match) > 0:
                    result['project_type_moj'] = lc_match.iloc[0].get('project_type', '')
            else:
                # Deal nie znaleziony - szukaj po nazwie
                deal_info, score, acc_name = find_deal_by_name(klient)
                if deal_info:
                    result['deal_id'] = deal_info['deal_id']
                    result['account_id'] = deal_info['account_id']
                    result['account_name_zoho'] = deal_info['account_name']
                    result['deal_stage'] = deal_info['deal_stage']
                    result['deal_owner'] = deal_info['deal_owner']
                    result['project_type_moj'] = deal_info['project_type']
                    result['link_deal'] = f"{ZOHO_BASE}/Potentials/{deal_info['deal_id']}"
                    if deal_info['account_id']:
                        result['link_account'] = f"{ZOHO_BASE}/Accounts/{deal_info['account_id']}"
                    result['match_method'] = f'LINK->DEAL brak, NAZWA ({int(score*100)}%)'
                    result['uwagi'] = f"Deal z linku nie istnieje, znaleziono po nazwie"
                else:
                    result['match_method'] = 'BŁĄD'
                    result['uwagi'] = f"Deal {helper['record_id']} nie istnieje, nie znaleziono po nazwie"
        
        # Dla innych typów linków - szukaj po nazwie
        else:
            deal_info, score, acc_name = find_deal_by_name(klient)
            if deal_info:
                result['deal_id'] = deal_info['deal_id']
                result['account_id'] = deal_info['account_id']
                result['account_name_zoho'] = deal_info['account_name']
                result['deal_stage'] = deal_info['deal_stage']
                result['deal_owner'] = deal_info['deal_owner']
                result['project_type_moj'] = deal_info['project_type']
                result['link_deal'] = f"{ZOHO_BASE}/Potentials/{deal_info['deal_id']}"
                if deal_info['account_id']:
                    result['link_account'] = f"{ZOHO_BASE}/Accounts/{deal_info['account_id']}"
                
                if score >= 0.9:
                    result['match_method'] = f'LINK->{helper["record_type"]}, NAZWA DOKŁADNA ({int(score*100)}%)'
                else:
                    result['match_method'] = f'LINK->{helper["record_type"]}, NAZWA ({int(score*100)}%)'
                    result['uwagi'] = f"Sprawdź: '{klient}' vs '{acc_name}'"
            else:
                result['match_method'] = f'LINK->{helper["record_type"]}, NIE ZNALEZIONO'
                result['uwagi'] = f"Link do {helper['record_type']}, Deal nie znaleziony po nazwie ({int(score*100)}%)"
    
    # Brak w danych pomocniczych - szukaj tylko po nazwie
    else:
        deal_info, score, acc_name = find_deal_by_name(klient)
        if deal_info:
            result['deal_id'] = deal_info['deal_id']
            result['account_id'] = deal_info['account_id']
            result['account_name_zoho'] = deal_info['account_name']
            result['deal_stage'] = deal_info['deal_stage']
            result['deal_owner'] = deal_info['deal_owner']
            result['project_type_moj'] = deal_info['project_type']
            result['link_deal'] = f"{ZOHO_BASE}/Potentials/{deal_info['deal_id']}"
            if deal_info['account_id']:
                result['link_account'] = f"{ZOHO_BASE}/Accounts/{deal_info['account_id']}"
            
            if score >= 0.9:
                result['match_method'] = f'TYLKO NAZWA DOKŁADNA ({int(score*100)}%)'
            else:
                result['match_method'] = f'TYLKO NAZWA ({int(score*100)}%)'
                result['uwagi'] = f"Brak w sty-wrz, znaleziono: '{acc_name}'"
        else:
            result['match_method'] = 'NIE ZNALEZIONO'
            result['uwagi'] = f"Brak w sty-wrz, nie znaleziono po nazwie ({int(score*100)}%)"
    
    results.append(result)

# ========== STATYSTYKI ==========
print("\n" + "="*60)
print("STATYSTYKI DOPASOWANIA")
print("="*60)

methods = {}
for r in results:
    m = r['match_method'].split(',')[0] if ',' in r['match_method'] else r['match_method'].split(' ')[0]
    methods[m] = methods.get(m, 0) + 1

for m, c in sorted(methods.items(), key=lambda x: -x[1]):
    print(f"  {m}: {c}")

with_deal = len([r for r in results if r['deal_id']])
print(f"\nRekordy z Deal ID: {with_deal} / {len(results)}")

# ========== EKSPORT DO XLSX ==========
wb = Workbook()
ws = wb.active
ws.title = "PreSales Umowy Kompletne"

headers = [
    'Lp', 'Klient', 'Produkt', 'PreSales', 'Miesiąc',
    'Typ źródła', 'Źródło szczegół',
    'Link źródłowy', 'Typ linku',
    'Link do Deal', 'Link do Account',
    'Deal ID', 'Account ID', 'Nazwa firmy Zoho',
    'Deal Stage', 'Deal Owner', 'Mój typ projektu',
    'Metoda dopasowania', 'Uwagi'
]

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
link_font = Font(color="0563C1", underline="single")
warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
manual_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for row_idx, row in enumerate(results):
    r = row_idx + 2
    
    ws.cell(row=r, column=1, value=row['lp'])
    ws.cell(row=r, column=2, value=row['klient'])
    ws.cell(row=r, column=3, value=row['produkt'])
    ws.cell(row=r, column=4, value=row['presales'])
    ws.cell(row=r, column=5, value=row['miesiac'])
    ws.cell(row=r, column=6, value=row['typ_zrodla'])
    ws.cell(row=r, column=7, value=row['zrodlo_szczegol'])
    
    if row['link_zrodlowy']:
        cell = ws.cell(row=r, column=8, value=row['link_zrodlowy'])
        cell.hyperlink = row['link_zrodlowy']
        cell.font = link_font
    
    ws.cell(row=r, column=9, value=row['link_type_zrodlowy'])
    
    if row['link_deal']:
        cell = ws.cell(row=r, column=10, value=row['link_deal'])
        cell.hyperlink = row['link_deal']
        cell.font = link_font
    
    if row['link_account']:
        cell = ws.cell(row=r, column=11, value=row['link_account'])
        cell.hyperlink = row['link_account']
        cell.font = link_font
    
    ws.cell(row=r, column=12, value=row['deal_id'])
    ws.cell(row=r, column=13, value=row['account_id'])
    ws.cell(row=r, column=14, value=row['account_name_zoho'])
    ws.cell(row=r, column=15, value=row['deal_stage'])
    ws.cell(row=r, column=16, value=row['deal_owner'])
    ws.cell(row=r, column=17, value=row['project_type_moj'])
    ws.cell(row=r, column=18, value=row['match_method'])
    ws.cell(row=r, column=19, value=row['uwagi'])
    
    # Kolorowanie
    method = row['match_method']
    if 'BEZPOŚREDNI' in method or 'DOKŁADNA' in method:
        fill = ok_fill
    elif 'TYLKO NAZWA' in method and 'DOKŁADNA' not in method:
        fill = manual_fill  # niebieski - brak w sty-wrz, ale znaleziono
    elif 'NIE ZNALEZIONO' in method or 'BŁĄD' in method:
        fill = warning_fill
    else:
        fill = partial_fill
    
    for col in range(1, 20):
        ws.cell(row=r, column=col).fill = fill

widths = [5, 35, 12, 25, 8, 12, 25, 50, 10, 50, 50, 22, 22, 40, 15, 20, 12, 35, 50]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = 'A2'

output = 'raw_data/presales_umowy_kompletne.xlsx'
wb.save(output)
print(f"\nZapisano: {output}")

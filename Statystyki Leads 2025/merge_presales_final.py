"""Scal plik źródłowy PreSales z dopasowaniem i przetłumacz Leads na Deale."""

import pandas as pd
import json
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from difflib import SequenceMatcher

ZOHO_BASE = "https://crm.zoho.eu/crm/org20101283812/tab"

def extract_id_from_link(link):
    """Wyciągnij ID i typ z linku Zoho."""
    if not link or pd.isna(link):
        return None, None
    link = str(link).strip()
    patterns = [
        (r'/tab/Potentials/(\d+)', 'Deal'),
        (r'/tab/Leads/(\d+)', 'Lead'),
        (r'/tab/Accounts/(\d+)', 'Account'),
        (r'/tab/Events/(\d+)', 'Event'),
    ]
    for pattern, obj_type in patterns:
        match = re.search(pattern, link)
        if match:
            return match.group(1), obj_type
    return None, None

def normalize(s):
    """Normalizuj nazwę do porównania."""
    if not s:
        return ''
    s = str(s).lower()
    s = re.sub(r'[^a-ząćęłńóśźż0-9\s]', ' ', s)  # zachowaj spacje
    s = ' '.join(s.split())  # usuń wielokrotne spacje
    return s

def get_key_words(s):
    """Wyciągnij kluczowe słowa (>2 znaki)."""
    norm = normalize(s)
    return set(w for w in norm.split() if len(w) > 2)

def similarity(a, b):
    """Oblicz podobieństwo - łącz różne metody."""
    norm_a = normalize(a).replace(' ', '')
    norm_b = normalize(b).replace(' ', '')
    
    # 1. Dokładne dopasowanie bez spacji
    if norm_a == norm_b:
        return 1.0
    
    # 2. Jeden zawiera się w drugim
    if norm_a in norm_b or norm_b in norm_a:
        return 0.95
    
    # 3. Słowa kluczowe
    words_a = get_key_words(a)
    words_b = get_key_words(b)
    if words_a and words_b:
        common = words_a & words_b
        if len(common) >= 2:  # co najmniej 2 wspólne słowa
            return 0.9
        if len(common) == 1 and len(words_a) <= 2:  # 1 słowo gdy krótka nazwa
            return 0.85
    
    # 4. SequenceMatcher jako fallback
    return SequenceMatcher(None, norm_a, norm_b).ratio()

# Wczytaj Deals z lifecycle
lifecycle = pd.read_csv('curated_data/lifecycle_2025.csv', dtype=str)
deals_df = lifecycle[lifecycle['deal_id'].notna()].copy()

# Wczytaj raw deals dla dodatkowych informacji
deals_raw = json.load(open('raw_data/deals_raw.json', 'r', encoding='utf-8'))
deals_map = {str(d.get('id')): d for d in deals_raw}

# Buduj mapę: account_name -> deal_id (wybierz najnowszy deal dla danego account)
account_to_deals = {}
for _, row in deals_df.iterrows():
    acc_name = normalize(row.get('account_name', ''))
    if acc_name:
        if acc_name not in account_to_deals:
            account_to_deals[acc_name] = []
        account_to_deals[acc_name].append({
            'deal_id': row['deal_id'],
            'account_id': row.get('account_id', ''),
            'account_name': row.get('account_name', ''),
            'deal_stage': row.get('deal_stage', ''),
            'deal_owner': row.get('deal_owner', ''),
            'deal_created': row.get('deal_stage_start_time', ''),
            'project_type': row.get('project_type', '')
        })

print(f"Unikalne nazwy firm w deals: {len(account_to_deals)}")

# Wczytaj plik źródłowy PreSales (styczeń-wrzesień)
source_data = []
with open('raw_data/sprzedaz_presales_source_sztyczen_wrzesien_umowy.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)  # skip header
    for row in reader:
        if row and row[0].strip() and row[0].strip() not in ['miesiąc podpisania umowy', '']:
            link = row[7].strip() if len(row) > 7 else ''
            record_id, record_type = extract_id_from_link(link)
            source_data.append({
                'miesiac': row[0].strip(),
                'klient': row[1].strip(),
                'produkt': row[2].strip() if len(row) > 2 else '',
                'owner': row[3].strip() if len(row) > 3 else '',
                'kwota_wdrozenie': row[5].strip() if len(row) > 5 else '',
                'kwota_abo': row[6].strip() if len(row) > 6 else '',
                'link_zrodlowy': link,
                'record_id': record_id,
                'record_type': record_type,
                'typ_zrodla': row[8].strip() if len(row) > 8 else '',
                'zrodlo_szczegol': row[9].strip() if len(row) > 9 else '',
                'presales': row[10].strip() if len(row) > 10 else ''
            })

print(f"Rekordy z pliku źródłowego: {len(source_data)}")

# Teraz dla każdego rekordu znajdź Deal
results = []
for idx, src in enumerate(source_data, 1):
    result = {
        'lp': idx,
        'miesiac': src['miesiac'],
        'klient': src['klient'],
        'produkt': src['produkt'],
        'owner_presales': src['presales'],
        'typ_zrodla': src['typ_zrodla'],
        'zrodlo_szczegol': src['zrodlo_szczegol'],
        'kwota_wdrozenie': src['kwota_wdrozenie'],
        'kwota_abo': src['kwota_abo'],
        'link_zrodlowy': src['link_zrodlowy'],
        'link_type_zrodlowy': src['record_type'],
        'deal_id': '',
        'account_id': '',
        'account_name_zoho': '',
        'deal_stage': '',
        'deal_owner': '',
        'project_type_moj': '',
        'link_deal': '',
        'link_account': '',
        'match_method': '',
        'uwagi': ''
    }
    
    # Jeśli link źródłowy to Deal - użyj go bezpośrednio
    if src['record_type'] == 'Deal' and src['record_id']:
        deal = deals_map.get(src['record_id'])
        if deal:
            result['deal_id'] = src['record_id']
            acc = deal.get('Account_Name', {})
            result['account_id'] = str(acc.get('id', '')) if isinstance(acc, dict) else ''
            result['account_name_zoho'] = acc.get('name', '') if isinstance(acc, dict) else str(acc)
            result['deal_stage'] = deal.get('Stage', '')
            owner = deal.get('Owner', {})
            result['deal_owner'] = owner.get('name', '') if isinstance(owner, dict) else str(owner)
            result['link_deal'] = f"{ZOHO_BASE}/Potentials/{src['record_id']}"
            if result['account_id']:
                result['link_account'] = f"{ZOHO_BASE}/Accounts/{result['account_id']}"
            result['match_method'] = 'BEZPOŚREDNI (Deal)'
            
            # Znajdź project_type z lifecycle
            lc_match = deals_df[deals_df['deal_id'] == src['record_id']]
            if len(lc_match) > 0:
                result['project_type_moj'] = lc_match.iloc[0].get('project_type', '')
        else:
            result['uwagi'] = f"Deal {src['record_id']} nie znaleziony w dumpie"
            result['match_method'] = 'BŁĄD'
    
    # Dla Leads, Events, Accounts - szukaj po nazwie firmy
    else:
        klient = src['klient']
        
        # Szukaj najlepszego dopasowania
        best_match = None
        best_score = 0
        best_acc_name = ''
        
        for acc_name_orig, deals_list in account_to_deals.items():
            # Znajdź oryginalną nazwę (nie znormalizowaną)
            orig_name = deals_list[0]['account_name'] if deals_list else ''
            score = similarity(klient, orig_name)
            if score > best_score:
                best_score = score
                best_match = deals_list
                best_acc_name = orig_name
        
        if best_match and best_score >= 0.5:
            # Wybierz deal z najwyższym stage (preferuj won)
            won_stages = ['Closed Won', 'Wdrożenie', 'Trial', 'Wdrożeni Klienci']
            won_deals = [d for d in best_match if d['deal_stage'] in won_stages]
            
            if won_deals:
                deal_info = won_deals[0]
            else:
                deal_info = best_match[0]
            
            result['deal_id'] = deal_info['deal_id']
            result['account_id'] = deal_info['account_id']
            result['account_name_zoho'] = deal_info['account_name']
            result['deal_stage'] = deal_info['deal_stage']
            result['deal_owner'] = deal_info['deal_owner']
            result['project_type_moj'] = deal_info['project_type']
            
            if deal_info['deal_id']:
                result['link_deal'] = f"{ZOHO_BASE}/Potentials/{deal_info['deal_id']}"
            if deal_info['account_id']:
                result['link_account'] = f"{ZOHO_BASE}/Accounts/{deal_info['account_id']}"
            
            if best_score >= 0.95:
                result['match_method'] = f'DOKŁADNE ({int(best_score*100)}%)'
            else:
                result['match_method'] = f'PRZYBLIŻONE ({int(best_score*100)}%)'
                result['uwagi'] = f"Sprawdź: '{src['klient']}' vs '{deal_info['account_name']}'"
        else:
            result['match_method'] = 'NIE ZNALEZIONO'
            result['uwagi'] = f"Brak dopasowania dla '{src['klient']}' (najlepsze: {int(best_score*100)}%)"
    
    results.append(result)

# Statystyki
methods = {}
for r in results:
    m = r['match_method'].split(' ')[0]
    methods[m] = methods.get(m, 0) + 1

print("\nMetody dopasowania:")
for m, c in sorted(methods.items(), key=lambda x: -x[1]):
    print(f"  {m}: {c}")

# Eksport do XLSX
wb = Workbook()
ws = wb.active
ws.title = "PreSales Umowy"

headers = [
    'Lp', 'Miesiąc', 'Klient', 'Produkt', 'PreSales', 
    'Typ źródła', 'Źródło szczegół', 'Kwota wdrożenie', 'Kwota abo',
    'Link źródłowy', 'Typ linku źródł.',
    'Link do Deal', 'Link do Account',
    'Deal ID', 'Account ID', 'Nazwa firmy Zoho',
    'Deal Stage', 'Deal Owner', 'Mój typ projektu',
    'Metoda dopasowania', 'Uwagi'
]

# Style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
link_font = Font(color="0563C1", underline="single")
warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

# Nagłówki
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Dane
for row_idx, row in enumerate(results):
    r = row_idx + 2
    
    ws.cell(row=r, column=1, value=row['lp'])
    ws.cell(row=r, column=2, value=row['miesiac'])
    ws.cell(row=r, column=3, value=row['klient'])
    ws.cell(row=r, column=4, value=row['produkt'])
    ws.cell(row=r, column=5, value=row['owner_presales'])
    ws.cell(row=r, column=6, value=row['typ_zrodla'])
    ws.cell(row=r, column=7, value=row['zrodlo_szczegol'])
    ws.cell(row=r, column=8, value=row['kwota_wdrozenie'])
    ws.cell(row=r, column=9, value=row['kwota_abo'])
    
    # Link źródłowy
    if row['link_zrodlowy']:
        cell = ws.cell(row=r, column=10, value=row['link_zrodlowy'])
        cell.hyperlink = row['link_zrodlowy']
        cell.font = link_font
    
    ws.cell(row=r, column=11, value=row['link_type_zrodlowy'])
    
    # Link do Deal
    if row['link_deal']:
        cell = ws.cell(row=r, column=12, value=row['link_deal'])
        cell.hyperlink = row['link_deal']
        cell.font = link_font
    
    # Link do Account
    if row['link_account']:
        cell = ws.cell(row=r, column=13, value=row['link_account'])
        cell.hyperlink = row['link_account']
        cell.font = link_font
    
    ws.cell(row=r, column=14, value=row['deal_id'])
    ws.cell(row=r, column=15, value=row['account_id'])
    ws.cell(row=r, column=16, value=row['account_name_zoho'])
    ws.cell(row=r, column=17, value=row['deal_stage'])
    ws.cell(row=r, column=18, value=row['deal_owner'])
    ws.cell(row=r, column=19, value=row['project_type_moj'])
    ws.cell(row=r, column=20, value=row['match_method'])
    ws.cell(row=r, column=21, value=row['uwagi'])
    
    # Kolorowanie
    method = row['match_method']
    if 'BEZPOŚREDNI' in method or 'DOKŁADNE' in method:
        for col in range(1, 22):
            ws.cell(row=r, column=col).fill = ok_fill
    elif 'PRZYBLIŻONE' in method:
        for col in range(1, 22):
            ws.cell(row=r, column=col).fill = partial_fill
    elif 'NIE ZNALEZIONO' in method or 'BŁĄD' in method:
        for col in range(1, 22):
            ws.cell(row=r, column=col).fill = warning_fill

# Szerokości kolumn
widths = [5, 8, 35, 12, 20, 12, 25, 15, 12, 50, 10, 50, 50, 22, 22, 40, 15, 20, 12, 20, 50]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = 'A2'

output = 'raw_data/presales_umowy_final.xlsx'
wb.save(output)
print(f"\nZapisano: {output}")

# Podsumowanie
with_deal = len([r for r in results if r['deal_id']])
print(f"Rekordy z Deal ID: {with_deal} / {len(results)}")

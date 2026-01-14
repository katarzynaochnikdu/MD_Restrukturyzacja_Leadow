"""
Porównanie lifecycle z oryginalnym Excelem.

Analizuje różnice między:
- lifecycle_2025.csv (nasze obliczenia)
- 202511 Podsumowanie Leadów i MLi_backup.xlsx (oryginał)
"""

import logging
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import pandas as pd

from config import MONTHS_2025, MONTH_NUM_TO_NAME, CURATED_DATA_DIR

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

LIFECYCLE_FILE = Path(CURATED_DATA_DIR) / "lifecycle_2025.csv"
EXCEL_DIR = Path(r"C:/Users/kochn/.cursor/Medidesk/Statystyki Leads 2025")


def find_backup_excel() -> Path:
    """Znajduje plik backup Excel."""
    files = [f for f in EXCEL_DIR.glob("*Podsumowanie*backup*") if not f.name.startswith("~")]
    if not files:
        raise FileNotFoundError("Nie znaleziono pliku backup Excel")
    return files[0]


def load_lifecycle() -> pd.DataFrame:
    """Wczytuje lifecycle."""
    return pd.read_csv(LIFECYCLE_FILE, dtype=str)


def parse_dt(s: str) -> Tuple[int, int]:
    """Wyciąga rok i miesiąc z ISO datetime string."""
    if not s or pd.isna(s):
        return None, None
    try:
        return int(s[:4]), int(s[5:7])
    except (ValueError, IndexError):
        return None, None


def extract_excel_monthly(ws, months_cols: range = range(5, 16)) -> Dict[str, int]:
    """Wyciąga sumy miesięczne z arkusza Excel."""
    months = MONTHS_2025[:len(list(months_cols))]
    totals = {}
    
    for col_idx, month in zip(months_cols, months):
        total = 0
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, col_idx).value
            if val and isinstance(val, (int, float)):
                total += int(val)
        totals[month] = total
    
    return totals


def calculate_lifecycle_monthly(df: pd.DataFrame, date_col: str, filter_fn=None) -> Dict[str, int]:
    """Oblicza sumy miesięczne z lifecycle."""
    if filter_fn:
        df = df[filter_fn(df)]
    
    totals = {m: 0 for m in MONTHS_2025}
    
    for _, row in df.iterrows():
        year, month = parse_dt(row.get(date_col, ""))
        if year == 2025 and month:
            month_name = MONTH_NUM_TO_NAME.get(month)
            if month_name:
                totals[month_name] += 1
    
    return totals


def print_comparison(title: str, lifecycle_totals: Dict[str, int], excel_totals: Dict[str, int]):
    """Wyświetla porównanie."""
    print(f"\n{'='*60}")
    print(f"{title}")
    print(f"{'='*60}")
    print(f"{'Miesiąc':<12} {'Lifecycle':>10} {'Excel':>10} {'Różnica':>10} {'%':>8}")
    print("-" * 52)
    
    lc_total = 0
    ex_total = 0
    
    for month in MONTHS_2025[:11]:  # Jan-Lis (bez Grudnia bo go nie ma w backup)
        lc = lifecycle_totals.get(month, 0)
        ex = excel_totals.get(month, 0)
        diff = lc - ex
        pct = (diff / ex * 100) if ex > 0 else 0
        
        print(f"{month:<12} {lc:>10} {ex:>10} {diff:>+10} {pct:>+7.1f}%")
        lc_total += lc
        ex_total += ex
    
    print("-" * 52)
    diff_total = lc_total - ex_total
    pct_total = (diff_total / ex_total * 100) if ex_total > 0 else 0
    print(f"{'SUMA':<12} {lc_total:>10} {ex_total:>10} {diff_total:>+10} {pct_total:>+7.1f}%")


def main():
    logger.info("Wczytywanie danych...")
    
    # Lifecycle
    df = load_lifecycle()
    logger.info(f"Lifecycle: {len(df)} wierszy")
    
    # Excel
    excel_path = find_backup_excel()
    logger.info(f"Excel: {excel_path.name}")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # === ANALIZA 1: MLe ===
    # W Excelu: arkusz "MLe" zawiera WSZYSTKIE Marketing Leads pogrupowane per źródło
    # Nasza definicja: ML bez konwersji na Leada
    
    ws_mle = wb["MLe"]
    excel_mle = extract_excel_monthly(ws_mle)
    
    # Nasze MLe = ML bez Lead (ml_id not empty AND lead_id empty)
    lifecycle_mle = calculate_lifecycle_monthly(
        df, 
        "ml_stage_start_time",
        lambda d: (d["ml_id"].notna() & (d["ml_id"] != "")) & 
                  (d["lead_id"].isna() | (d["lead_id"] == ""))
    )
    
    # Wszystkie ML (dla porównania)
    lifecycle_all_ml = calculate_lifecycle_monthly(
        df,
        "ml_stage_start_time", 
        lambda d: d["ml_id"].notna() & (d["ml_id"] != "")
    )
    
    print_comparison("MLe (ML bez konwersji na Lead) vs Excel MLe", lifecycle_mle, excel_mle)
    print_comparison("WSZYSTKIE ML vs Excel MLe", lifecycle_all_ml, excel_mle)
    
    # === ANALIZA 2: MLe+ Leady ===
    ws_leads = wb["MLe+ Leady"]
    excel_leads = extract_excel_monthly(ws_leads)
    
    # Nasze Leady = wszystkie z lead_id
    lifecycle_leads = calculate_lifecycle_monthly(
        df,
        "lead_stage_start_time",
        lambda d: d["lead_id"].notna() & (d["lead_id"] != "")
    )
    
    print_comparison("Wszystkie Leady vs Excel MLe+ Leady", lifecycle_leads, excel_leads)
    
    # === PODSUMOWANIE ===
    print("\n" + "="*60)
    print("PODSUMOWANIE")
    print("="*60)
    
    # Statystyki z lifecycle
    total_cycles = len(df)
    with_ml = len(df[(df["ml_id"].notna()) & (df["ml_id"] != "")])
    with_lead = len(df[(df["lead_id"].notna()) & (df["lead_id"] != "")])
    with_deal = len(df[(df["deal_id"].notna()) & (df["deal_id"] != "")])
    ml_only = len(df[
        (df["ml_id"].notna() & (df["ml_id"] != "")) & 
        (df["lead_id"].isna() | (df["lead_id"] == ""))
    ])
    ml_with_lead = with_ml - ml_only
    
    print(f"\nZ lifecycle_2025.csv (łącznie {total_cycles} cykli):")
    print(f"  - Cykli z ML: {with_ml}")
    print(f"    - ML bez Lead (MLe): {ml_only}")
    print(f"    - ML z Lead: {ml_with_lead}")
    print(f"  - Cykli z Lead: {with_lead}")
    print(f"  - Cykli z Deal: {with_deal}")
    
    print("\nWNIOSKI:")
    print("  1. Excel 'MLe' wydaje się zawierać WSZYSTKIE Marketing Leads")
    print("     (nie tylko te bez konwersji na Lead)")
    print("  2. Nasza definicja 'MLe' = ML bez konwersji na Lead")
    print("  3. Różnice mogą wynikać z:")
    print("     - Innej definicji w oryginalnym Excelu")
    print("     - Rekordów dodanych/usuniętych od czasu utworzenia Excela")
    print("     - Różnic w mapowaniu źródeł")


if __name__ == "__main__":
    main()

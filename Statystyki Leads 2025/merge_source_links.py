"""Połącz linki z pliku źródłowego PreSales z moim dopasowaniem."""

import pandas as pd
import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ZOHO_BASE = "https://crm.zoho.eu/crm/org20101283812/tab"

# Wczytaj plik źródłowy (styczeń-wrzesień z pełnymi linkami)
source_jan_sep = {}
with open('raw_data/sprzedaz_presales_source_sztyczen_wrzesien_umowy.csv', 'r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=';')
    next(reader)
    for row in reader:
        if row and len(row) > 1 and row[1].strip():
            klient = row[1].strip().lower().replace('-', '')
            link = row[7].strip() if len(row) > 7 else ''
            owner = row[3].strip() if len(row) > 3 else ''
            miesiac = row[0].strip() if row[0].strip() else ''
            source_jan_sep[klient] = {
                'link': link,
                'owner': owner,
                'miesiac': miesiac
            }

print(f"Linki z pliku źródłowego (sty-wrz): {len(source_jan_sep)}")

# Wczytaj moje dopasowanie
my_df = pd.read_csv('raw_data/presales_deals_matching.csv', dtype=str)

# Wczytaj lifecycle
lifecycle = pd.read_csv('curated_data/lifecycle_2025.csv', dtype={'deal_id': str, 'account_id': str})
deal_to_account = dict(zip(lifecycle['deal_id'].dropna(), lifecycle['account_id'].dropna()))

# Aktualizuj dopasowanie
results = []
for _, row in my_df.iterrows():
    klient_key = row['ps_klient'].lower().replace('-', '')
    
    result = dict(row)
    
    # Sprawdź czy mamy link z pliku źródłowego
    if klient_key in source_jan_sep:
        src = source_jan_sep[klient_key]
        result['link_zrodlowy'] = src['link']
        result['owner_zrodlowy'] = src['owner']
        result['miesiac_zrodlowy'] = src['miesiac']
    else:
        result['link_zrodlowy'] = ''
        result['owner_zrodlowy'] = ''
        result['miesiac_zrodlowy'] = ''
    
    # Ustaw ostateczny link
    if result['link_zrodlowy']:
        result['link_final'] = result['link_zrodlowy']
    elif pd.notna(result.get('deal_id')) and result.get('deal_id'):
        result['link_final'] = f"{ZOHO_BASE}/Potentials/{result['deal_id']}"
    else:
        result['link_final'] = ''
    
    results.append(result)

results_df = pd.DataFrame(results)

# Eksport do XLSX
wb = Workbook()
ws = wb.active
ws.title = "PreSales Dopasowanie"

headers = [
    'Lp', 'Klient PreSales', 'Produkt', 'PreSales', 'Miesiąc',
    'Link do rekordu', 'Typ linku',
    'Mój Deal ID', 'Mój Account ID', 'Nazwa firmy Zoho',
    'Mój typ projektu', 'Deal Owner', 'Jakość dopasowania', 'Uwagi'
]

# Style
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
link_font = Font(color="0563C1", underline="single")
warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Nagłówki
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Dane
for row_idx, row in results_df.iterrows():
    r = row_idx + 2
    
    # Określ typ linku
    link = row.get('link_final', '')
    if '/Leads/' in link:
        link_type = 'Lead'
    elif '/Potentials/' in link:
        link_type = 'Deal'
    elif '/Events/' in link:
        link_type = 'Event'
    elif '/Accounts/' in link:
        link_type = 'Account'
    else:
        link_type = ''
    
    ws.cell(row=r, column=1, value=row.get('lp', ''))
    ws.cell(row=r, column=2, value=row.get('ps_klient', ''))
    ws.cell(row=r, column=3, value=row.get('ps_produkt', ''))
    ws.cell(row=r, column=4, value=row.get('ps_presales', ''))
    ws.cell(row=r, column=5, value=row.get('miesiac_zrodlowy', ''))
    
    # Link
    if link:
        cell = ws.cell(row=r, column=6, value=link)
        cell.hyperlink = link
        cell.font = link_font
    
    ws.cell(row=r, column=7, value=link_type)
    ws.cell(row=r, column=8, value=row.get('deal_id', ''))
    ws.cell(row=r, column=9, value=row.get('account_id', ''))
    ws.cell(row=r, column=10, value=row.get('account_name_zoho', ''))
    ws.cell(row=r, column=11, value=row.get('moj_project_type', ''))
    ws.cell(row=r, column=12, value=row.get('deal_owner', ''))
    ws.cell(row=r, column=13, value=row.get('match_quality', ''))
    ws.cell(row=r, column=14, value=row.get('uwagi', ''))
    
    # Kolorowanie
    uwagi = str(row.get('uwagi', ''))
    match_q = str(row.get('match_quality', ''))
    
    if 'RÓŻNICA' in uwagi or 'SPRAWDŹ' in uwagi or match_q == 'NIE ZNALEZIONO':
        for col in range(1, 15):
            ws.cell(row=r, column=col).fill = warning_fill
    elif 'DOKŁADNE' in match_q or '90%' in match_q:
        for col in range(1, 15):
            ws.cell(row=r, column=col).fill = ok_fill

# Szerokości kolumn
widths = [5, 30, 12, 20, 8, 70, 8, 22, 22, 45, 12, 20, 18, 35]
for col, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = 'A2'

output = 'raw_data/presales_deals_final.xlsx'
wb.save(output)
print(f"Zapisano: {output}")

# Statystyki
with_link = len([r for r in results if r.get('link_final')])
print(f"Rekordy z linkiem: {with_link} / {len(results)}")
